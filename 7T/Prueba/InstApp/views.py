from io import BytesIO
from types import SimpleNamespace

from django.contrib import messages
from django.db import IntegrityError, transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from GuardaApp.views import (
    _filtrar_ambientes,
    _filtrar_incidentes,
    _filtrar_minutas,
    _filtrar_traslados,
    _q_param,
    _safe_q_part,
)
from LoginApp.models import (
    Ambiente,
    RegistroIncidente,
    RegistroMinuta,
    TrasladoRecurso,
    UserRol,
    Usuario,
)
from Prueba.report_utils import landscape_pdf_response


def _fecha_hoy():
    return timezone.localdate().isoformat()


def _asistencia_vacia(asistencia_id=None):
    return SimpleNamespace(
        id=asistencia_id,
        aprendiz_id=None,
        instructor_id=None,
        jornada_id=None,
        estado="S",
        fecha=timezone.localdate(),
    )


def _traslado_vacio(traslado_id=None):
    return SimpleNamespace(
        id_traslado=traslado_id,
        recurso_id=None,
        ambiente_origen_id=None,
        ambiente_destino_id=None,
        fecha=timezone.localtime(),
        observacion="",
    )


def _incidente_vacio(incidente_id=None):
    return SimpleNamespace(
        id=incidente_id,
        descripcion="",
        fecha=timezone.localdate(),
        hora="08:00",
        ambiente=SimpleNamespace(id=None),
        tipo_incidente=SimpleNamespace(id=None),
    )


def _mensaje_pendiente(request, texto):
    messages.info(request, texto)


def _no_cache_response(response: HttpResponse) -> HttpResponse:
    response["Cache-Control"] = "no-cache, no-store, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


def _sesion_instructor(request):
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return None, _no_cache_response(redirect("login"))
    usuario = Usuario.objects.filter(pk=usuario_id).first()
    if not usuario:
        return None, _no_cache_response(redirect("login"))
    ur = UserRol.objects.filter(id_usuario=usuario).select_related("id_rol").first()
    rol = (ur.id_rol.nombre_rol or "").strip().lower() if ur else ""
    if rol != "instructor":
        messages.error(request, "No tienes permisos para acceder a esta sección.")
        return None, _no_cache_response(redirect("login"))
    return usuario, None


def instructor_index(request):
    return render(request, "instindex.html")


def inicio_instructor(request):
    return instructor_index(request)


def index_instructor(request):
    return instructor_index(request)


def mis_fichas(request):
    context = {"fichas": []}
    return render(request, "misFichas.html", context)


def ver_aprendices(request, ficha_id):
    context = {
        "ficha_id": ficha_id,
        "ficha": SimpleNamespace(id=ficha_id, num_ficha=ficha_id),
        "aprendices": [],
    }
    return render(request, "aprendicesFicha.html", context)


def listar_asistencias(request):
    return render(request, "asistencias/listarAsistencias.html", {"asistencias": []})


def listar_asistencia(request):
    return listar_asistencias(request)


def registrar_asistencia(request):
    if request.method == "POST":
        _mensaje_pendiente(
            request,
            "La logica para guardar asistencias se implementara despues.",
        )
        return redirect("instructor_asistencias")

    context = {
        "aprendices": [],
        "instructores": [],
        "jornadas": [],
        "today": _fecha_hoy(),
    }
    return render(request, "asistencias/formAsistencia.html", context)


def editar_asistencia(request, asistencia_id):
    if request.method == "POST":
        _mensaje_pendiente(
            request,
            f"La actualizacion de la asistencia {asistencia_id} queda pendiente de implementar.",
        )
        return redirect("instructor_asistencias")

    context = {
        "asistencia": _asistencia_vacia(asistencia_id),
        "aprendices": [],
        "instructores": [],
        "jornadas": [],
    }
    return render(request, "asistencias/editarAsistencia.html", context)


def eliminar_asistencia(request, asistencia_id):
    _mensaje_pendiente(
        request,
        f"La eliminacion de la asistencia {asistencia_id} se implementara luego.",
    )
    return redirect("instructor_asistencias")


def exportar_pdf(request):
    _mensaje_pendiente(request, "La exportacion de asistencias a PDF aun no esta implementada.")
    return redirect("instructor_asistencias")


def exportar_excel(request):
    _mensaje_pendiente(request, "La exportacion de asistencias a Excel aun no esta implementada.")
    return redirect("instructor_asistencias")


def listar_minutas(request):
    q = _q_param(request)
    qs = (
        RegistroMinuta.objects.select_related(
            "ambiente",
            "guarda_seguridad_usuario_id_usuario__usuario_id_usuario",
            "responsable__usuario_id_usuario",
        )
        .all()
        .order_by("-fecha_hora_recibo")
    )
    qs = _filtrar_minutas(qs, q)
    minutas = []
    for m in qs:
        gu = m.guarda_seguridad_usuario_id_usuario.usuario_id_usuario
        guarda_nombre = " ".join(filter(None, [gu.p_nombre, gu.p_apellido])).strip()
        ru = m.responsable.usuario_id_usuario
        responsable_nombre = " ".join(filter(None, [ru.p_nombre, ru.p_apellido])).strip()
        minutas.append(
            {
                "id_minuta": m.id_minuta,
                "ambiente_id": m.ambiente_id,
                "ambiente_nombre": str(m.ambiente.num_ambiente) if m.ambiente_id else None,
                "guarda_id": m.guarda_seguridad_usuario_id_usuario_id,
                "guarda_nombre": guarda_nombre or None,
                "responsable_id": m.responsable_id,
                "responsable_nombre": responsable_nombre or None,
                "fecha_recibo": m.fecha_hora_recibo,
                "fecha_entrega": m.fecha_hora_entrega,
                "estado": m.estado or "",
                "novedad": m.novedad or "",
            }
        )
    return render(request, "consultarMinutas.html", {"minutas": minutas, "q": q})


def consultar_minutas(request):
    return listar_minutas(request)


def _minutas_inst_qs(request):
    q = _q_param(request)
    qs = (
        RegistroMinuta.objects.select_related(
            "ambiente",
            "guarda_seguridad_usuario_id_usuario__usuario_id_usuario",
            "responsable__usuario_id_usuario",
        )
        .all()
        .order_by("-fecha_hora_recibo")
    )
    return _filtrar_minutas(qs, q), q


def exportar_minutas_pdf(request):
    minutas, q = _minutas_inst_qs(request)
    headers = ["ID", "Ambiente", "Recibo", "Entrega", "Estado", "Novedad", "Descripción"]
    rows = []
    for m in minutas:
        rows.append(
            [
                str(m.id_minuta),
                str(m.ambiente.num_ambiente) if m.ambiente_id else "",
                m.fecha_hora_recibo.strftime("%Y-%m-%d %H:%M") if m.fecha_hora_recibo else "",
                m.fecha_hora_entrega.strftime("%Y-%m-%d %H:%M") if m.fecha_hora_entrega else "",
                m.estado or "",
                m.novedad or "",
                m.descripcion_min or "",
            ]
        )
    return landscape_pdf_response(
        "Reporte de Minutas",
        q,
        headers,
        rows,
        [0.05, 0.07, 0.11, 0.11, 0.09, 0.22, 0.35],
        f"minutas_instructor_{_safe_q_part(q)}.pdf",
    )


def exportar_minutas_excel(request):
    minutas, q = _minutas_inst_qs(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Minutas"
    headers = ["ID", "Ambiente", "Recibo", "Entrega", "Estado", "Novedad", "Descripción"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for m in minutas:
        ws.append(
            [
                m.id_minuta,
                m.ambiente.num_ambiente if m.ambiente_id else "",
                m.fecha_hora_recibo.strftime("%Y-%m-%d %H:%M:%S") if m.fecha_hora_recibo else "",
                m.fecha_hora_entrega.strftime("%Y-%m-%d %H:%M:%S") if m.fecha_hora_entrega else "",
                m.estado or "",
                m.novedad or "",
                m.descripcion_min or "",
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="minutas_instructor_{_safe_q_part(q)}.xlsx"'
    return resp


def listar_incidentes(request):
    q = _q_param(request)
    qs = (
        RegistroIncidente.objects.select_related("ambiente", "tipo_inc", "usuario_id_usuario")
        .all()
        .order_by("-fecha_incidente", "-hora_incidente")
    )
    qs = _filtrar_incidentes(qs, q)
    incidentes = []
    for i in qs:
        rn = f"{i.usuario_id_usuario.p_nombre} {i.usuario_id_usuario.p_apellido}".strip()
        incidentes.append(
            SimpleNamespace(
                id=i.id_incidente,
                descripcion=i.descripcion or "",
                fecha=i.fecha_incidente,
                hora=i.hora_incidente,
                ambiente=SimpleNamespace(
                    nombre=str(i.ambiente.num_ambiente) if i.ambiente_id else None,
                ),
                tipo_incidente=SimpleNamespace(
                    nombre=i.tipo_inc.tipo_incidente if i.tipo_inc_id else None,
                ),
                reportador_nombre=rn or None,
            )
        )
    return render(request, "listarIncidentes.html", {"incidentes": incidentes, "q": q})


def crear_incidente(request):
    if request.method == "POST":
        _mensaje_pendiente(
            request,
            "La logica para guardar incidentes se implementara despues.",
        )
        return redirect("listar_incidentes")

    context = {
        "incidente": _incidente_vacio(),
        "ambientes": [],
        "tipos": [],
        "modo_edicion": False,
        "hoy": _fecha_hoy(),
    }
    return render(request, "formIncidente.html", context)


def form_incidente(request):
    return crear_incidente(request)


def editar_incidente(request, incidente_id):
    if request.method == "POST":
        _mensaje_pendiente(
            request,
            f"La actualizacion del incidente {incidente_id} queda pendiente de implementar.",
        )
        return redirect("listar_incidentes")

    context = {
        "incidente": _incidente_vacio(incidente_id),
        "ambientes": [],
        "tipos": [],
        "modo_edicion": True,
        "hoy": _fecha_hoy(),
    }
    return render(request, "formIncidente.html", context)


def eliminar_incidente(request, incidente_id):
    _mensaje_pendiente(
        request,
        f"La eliminacion del incidente {incidente_id} se implementara luego.",
    )
    return redirect("listar_incidentes")


def _incidentes_inst_qs(request):
    q = _q_param(request)
    qs = (
        RegistroIncidente.objects.select_related("ambiente", "tipo_inc", "usuario_id_usuario")
        .all()
        .order_by("-fecha_incidente", "-hora_incidente")
    )
    return _filtrar_incidentes(qs, q), q


def exportar_incidentes_pdf(request):
    incidentes, q = _incidentes_inst_qs(request)
    headers = ["ID", "Fecha", "Hora", "Ambiente", "Tipo", "Descripción", "Usuario"]
    rows = []
    for i in incidentes:
        usuario = f"{i.usuario_id_usuario.p_nombre} {i.usuario_id_usuario.p_apellido}".strip()
        rows.append(
            [
                str(i.id_incidente),
                i.fecha_incidente.isoformat() if i.fecha_incidente else "",
                str(i.hora_incidente) if i.hora_incidente else "",
                str(i.ambiente.num_ambiente) if i.ambiente_id else "",
                i.tipo_inc.tipo_incidente if i.tipo_inc_id else "",
                i.descripcion or "",
                usuario,
            ]
        )
    return landscape_pdf_response(
        "Reporte de Incidentes",
        q,
        headers,
        rows,
        [0.06, 0.09, 0.07, 0.08, 0.10, 0.35, 0.25],
        f"incidentes_instructor_{_safe_q_part(q)}.pdf",
    )


def exportar_incidentes_excel(request):
    incidentes, q = _incidentes_inst_qs(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Incidentes"
    headers = ["ID", "Fecha", "Hora", "Ambiente", "Tipo", "Descripción", "Usuario"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i in incidentes:
        usuario = f"{i.usuario_id_usuario.p_nombre} {i.usuario_id_usuario.p_apellido}".strip()
        ws.append(
            [
                i.id_incidente,
                i.fecha_incidente.isoformat() if i.fecha_incidente else "",
                str(i.hora_incidente) if i.hora_incidente else "",
                i.ambiente.num_ambiente if i.ambiente_id else "",
                i.tipo_inc.tipo_incidente if i.tipo_inc_id else "",
                i.descripcion or "",
                usuario,
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="incidentes_instructor_{_safe_q_part(q)}.xlsx"'
    return resp


def listar_traslados(request):
    q = _q_param(request)
    qs = TrasladoRecurso.objects.select_related("recurso", "ambiente_origen").all().order_by(
        "-fecha_traslado"
    )
    qs = _filtrar_traslados(qs, q)
    traslados = []
    for t in qs:
        traslados.append(
            {
                "id_traslado": t.id_traslado,
                "recurso_nombre": t.recurso.nombre_recurso if t.recurso_id else None,
                "recurso_id": t.recurso_id,
                "ambiente_origen_nombre": str(t.ambiente_origen.num_ambiente)
                if t.ambiente_origen_id
                else None,
                "ambiente_origen": t.ambiente_origen_id,
                "ambiente_destino_nombre": str(t.ambiente_destino)
                if t.ambiente_destino is not None
                else None,
                "ambiente_destino": t.ambiente_destino,
                "fecha_traslado": t.fecha_traslado,
                "observacion": t.observacion or "",
            }
        )
    return render(request, "traslados/listarTraslados.html", {"traslados": traslados, "q": q})


def form_traslado(request):
    if request.method == "POST":
        _mensaje_pendiente(
            request,
            "La logica para guardar traslados se implementara despues.",
        )
        return redirect("listar_traslados")

    context = {
        "traslado": _traslado_vacio(),
        "recursos": [],
        "ambientes": [],
        "modo_edicion": False,
    }
    return render(request, "traslados/formTraslado.html", context)


def editar_traslado(request, traslado_id):
    if request.method == "POST":
        _mensaje_pendiente(
            request,
            f"La actualizacion del traslado {traslado_id} queda pendiente de implementar.",
        )
        return redirect("listar_traslados")

    context = {
        "traslado": _traslado_vacio(traslado_id),
        "recursos": [],
        "ambientes": [],
        "modo_edicion": True,
    }
    return render(request, "traslados/formTraslado.html", context)


def eliminar_traslado(request, traslado_id):
    _mensaje_pendiente(
        request,
        f"La eliminacion del traslado {traslado_id} se implementara luego.",
    )
    return redirect("listar_traslados")


def _traslados_inst_qs(request):
    q = _q_param(request)
    qs = TrasladoRecurso.objects.select_related("recurso", "ambiente_origen").all().order_by(
        "-fecha_traslado"
    )
    return _filtrar_traslados(qs, q), q


def exportar_traslados_pdf(request):
    traslados, q = _traslados_inst_qs(request)
    headers = ["ID", "Recurso", "Serial", "Origen", "Destino", "Fecha", "Observación"]
    rows = []
    for t in traslados:
        rows.append(
            [
                str(t.id_traslado),
                t.recurso.nombre_recurso if t.recurso_id else "",
                t.recurso.serial_recurso if t.recurso_id else "",
                str(t.ambiente_origen.num_ambiente) if t.ambiente_origen_id else "",
                str(t.ambiente_destino or ""),
                t.fecha_traslado.strftime("%Y-%m-%d %H:%M:%S") if t.fecha_traslado else "",
                t.observacion or "",
            ]
        )
    return landscape_pdf_response(
        "Reporte de Traslados",
        q,
        headers,
        rows,
        [0.07, 0.18, 0.14, 0.10, 0.10, 0.14, 0.27],
        f"traslados_instructor_{_safe_q_part(q)}.pdf",
    )


def exportar_traslados_excel(request):
    traslados, q = _traslados_inst_qs(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Traslados"
    headers = ["ID", "Recurso", "Serial", "Origen", "Destino", "Fecha", "Observación"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for t in traslados:
        ws.append(
            [
                t.id_traslado,
                t.recurso.nombre_recurso if t.recurso_id else "",
                t.recurso.serial_recurso if t.recurso_id else "",
                t.ambiente_origen.num_ambiente if t.ambiente_origen_id else "",
                t.ambiente_destino or "",
                t.fecha_traslado.strftime("%Y-%m-%d %H:%M:%S") if t.fecha_traslado else "",
                t.observacion or "",
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="traslados_instructor_{_safe_q_part(q)}.xlsx"'
    return resp


def _ambientes_inst_list(request):
    q = _q_param(request)
    qs = Ambiente.objects.all().order_by("id_ambiente")
    qs = _filtrar_ambientes(qs, q)
    ambientes = []
    total_disponibles = 0
    total_ocupados = 0
    for a in qs:
        est = (a.estado or "").strip()
        if est in ("Disponible", "Operativo"):
            total_disponibles += 1
        elif est in ("Ocupado", "En Uso"):
            total_ocupados += 1
        ambientes.append(
            {
                "id_ambiente": a.id_ambiente,
                "numero": a.num_ambiente,
                "capacidad": a.capacidad,
                "tipo": a.tipo_ambiente or "",
                "estado": est or "—",
            }
        )
    return ambientes, total_disponibles, total_ocupados, q


def listar_ambientes(request):
    ambientes, total_disponibles, total_ocupados, q = _ambientes_inst_list(request)
    return render(
        request,
        "ambientes.html",
        {
            "ambientes": ambientes,
            "total_disponibles": total_disponibles,
            "total_ocupados": total_ocupados,
            "total_ambientes": len(ambientes),
            "q": q,
        },
    )


def consultar_ambientes(request):
    ambientes, _, _, q = _ambientes_inst_list(request)
    return render(request, "consultarAmbientes.html", {"ambientes": ambientes, "q": q})


def exportar_ambientes_pdf(request):
    q = _q_param(request)
    ambientes = _filtrar_ambientes(Ambiente.objects.all(), q).order_by("id_ambiente")
    headers = ["ID", "Número", "Capacidad", "Tipo", "Estado"]
    rows = []
    for a in ambientes:
        rows.append(
            [
                str(a.id_ambiente),
                str(a.num_ambiente),
                str(a.capacidad),
                a.tipo_ambiente or "",
                a.estado or "",
            ]
        )
    return landscape_pdf_response(
        "Reporte de Ambientes",
        q,
        headers,
        rows,
        [0.12, 0.14, 0.14, 0.35, 0.25],
        f"ambientes_instructor_{_safe_q_part(q)}.pdf",
    )


def exportar_ambientes_excel(request):
    q = _q_param(request)
    ambientes = _filtrar_ambientes(Ambiente.objects.all(), q).order_by("id_ambiente")
    wb = Workbook()
    ws = wb.active
    ws.title = "Ambientes"
    headers = ["ID", "Número", "Capacidad", "Tipo", "Estado"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for a in ambientes:
        ws.append(
            [
                a.id_ambiente,
                a.num_ambiente,
                a.capacidad,
                a.tipo_ambiente or "",
                a.estado or "",
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="ambientes_instructor_{_safe_q_part(q)}.xlsx"'
    return resp


def perfil(request):
    usuario, denied = _sesion_instructor(request)
    if denied:
        return denied

    user_rol = UserRol.objects.filter(id_usuario=usuario).select_related("id_rol").first()
    rol_nombre = (user_rol.id_rol.nombre_rol or "").strip() if user_rol else "Instructor"

    if request.method == "POST":
        usuario.p_nombre = (request.POST.get("p_nombre") or "").strip()
        usuario.s_nombre = (request.POST.get("s_nombre") or "").strip() or None
        usuario.p_apellido = (request.POST.get("p_apellido") or "").strip()
        usuario.s_apellido = (request.POST.get("s_apellido") or "").strip() or None
        usuario.correo = (request.POST.get("correo") or "").strip()
        try:
            with transaction.atomic():
                usuario.save()
            messages.success(request, "Perfil actualizado correctamente.")
            return _no_cache_response(redirect("instructor_perfil"))
        except IntegrityError:
            messages.error(request, "No se pudo actualizar el perfil.")

    return _no_cache_response(
        render(
            request,
            "inst/perfil.html",
            {"perfil": usuario, "rol_nombre": rol_nombre},
        )
    )
