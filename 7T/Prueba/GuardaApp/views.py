from datetime import datetime
from io import BytesIO
import json
import re

from django.contrib import messages
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST
from django.contrib.auth import logout

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from Prueba.report_utils import landscape_pdf_response, q_param as _q_param, safe_q_part as _safe_q_part

from LoginApp.models import (
    Ambiente,
    GuardaSeguridad,
    Instructor,
    Recursos,
    RegistroIncidente,
    RegistroMinuta,
    Rol,
    TrasladoRecurso,
    UserRol,
    Usuario,
)


def _no_cache(response: HttpResponse) -> HttpResponse:
    response["Cache-Control"] = "no-cache, no-store, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


def _usuario_sesion(request):
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return None
    return Usuario.objects.filter(pk=usuario_id).first()


def _es_guarda(usuario: Usuario) -> bool:
    rol = (
        UserRol.objects.select_related("id_rol")
        .filter(id_usuario=usuario)
        .first()
    )
    if not rol:
        return False
    return rol.id_rol.nombre_rol.strip().lower() == "guarda de seguridad"


def _acceso_guarda_o_login(request):
    usuario = _usuario_sesion(request)
    if not usuario:
        return None, _no_cache(redirect("login"))
    if not _es_guarda(usuario):
        messages.error(request, "No tienes permisos para entrar al módulo de guarda.")
        return None, _no_cache(redirect("login"))
    return usuario, None


def _render_guarda(request, template, context=None):
    usuario, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    data = {
        "usuario_actual": usuario,
    }
    if context:
        data.update(context)
    return _no_cache(render(request, template, data))


def _replace_redirect(url_name: str, kwargs=None) -> HttpResponse:
    """
    Reemplaza la navegación (no hace PUSH al historial), para que el botón `Atrás`
    no devuelva al formulario que acabas de enviar.
    """
    kwargs = kwargs or {}
    url = reverse(url_name, kwargs=kwargs)
    payload_url = json.dumps(url)
    html = f"""
<!doctype html>
<html lang="es">
<head>
    <meta charset="utf-8">
    <meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
    <meta http-equiv="Pragma" content="no-cache">
    <meta http-equiv="Expires" content="0">
</head>
<body>
    <script>
        window.location.replace({payload_url});
    </script>
</body>
</html>
"""
    return _no_cache(HttpResponse(html))


def _maybe_int(value: str):
    if value and re.fullmatch(r"\d+", str(value).strip()):
        return int(value)
    return None


def _filtrar_minutas(minutas_qs, q: str):
    if not q:
        return minutas_qs

    numero = _maybe_int(q)
    conds = (
        Q(estado__icontains=q)
        | Q(novedad__icontains=q)
        | Q(descripcion_min__icontains=q)
        | Q(responsable__usuario_id_usuario__p_nombre__icontains=q)
        | Q(responsable__usuario_id_usuario__p_apellido__icontains=q)
    )
    if numero is not None:
        conds |= Q(id_minuta=numero) | Q(ambiente__num_ambiente=numero)
    return minutas_qs.filter(conds)


def _filtrar_incidentes(incidentes_qs, q: str):
    if not q:
        return incidentes_qs

    numero = _maybe_int(q)
    conds = (
        Q(descripcion__icontains=q)
        | Q(tipo_inc__tipo_incidente__icontains=q)
        | Q(usuario_id_usuario__p_nombre__icontains=q)
        | Q(usuario_id_usuario__p_apellido__icontains=q)
    )

    # Soporta búsqueda por fecha/hora en formato ISO (YYYY-MM-DD / HH:MM)
    try:
        fecha = datetime.strptime(q, "%Y-%m-%d").date()
        conds |= Q(fecha_incidente=fecha)
    except ValueError:
        pass
    try:
        hora = datetime.strptime(q, "%H:%M").time()
        conds |= Q(hora_incidente=hora)
    except ValueError:
        pass

    if numero is not None:
        conds |= Q(id_incidente=numero) | Q(ambiente__num_ambiente=numero)

    return incidentes_qs.filter(conds)


def _filtrar_traslados(traslados_qs, q: str):
    if not q:
        return traslados_qs

    numero = _maybe_int(q)
    conds = (
        Q(observacion__icontains=q)
        | Q(recurso__nombre_recurso__icontains=q)
        | Q(recurso__serial_recurso__icontains=q)
    )
    if numero is not None:
        conds |= (
            Q(id_traslado=numero)
            | Q(ambiente_origen__num_ambiente=numero)
            | Q(ambiente_destino=numero)
        )
    return traslados_qs.filter(conds)


def _filtrar_ambientes(ambientes_qs, q: str):
    if not q:
        return ambientes_qs

    numero = _maybe_int(q)
    conds = Q(tipo_ambiente__icontains=q) | Q(estado__icontains=q)
    if numero is not None:
        conds |= (
            Q(id_ambiente=numero)
            | Q(num_ambiente=numero)
            | Q(capacidad=numero)
        )
    return ambientes_qs.filter(conds)


@never_cache
def guarda_index(request):
    usuario, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    return _render_guarda(request, "guarda/dashboard.html")


@never_cache
def cerrar_sesion_guarda(request):
    logout(request)
    response = redirect('login') 
    response["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response

@never_cache
def listar_minutas(request):
    q = _q_param(request)
    minutas = (
        RegistroMinuta.objects.select_related(
            "ambiente",
            "guarda_seguridad_usuario_id_usuario__usuario_id_usuario",
            "responsable__usuario_id_usuario",
        )
        .all()
        .order_by("-fecha_hora_recibo")
    )
    minutas = _filtrar_minutas(minutas, q)
    return _render_guarda(request, "guarda/minutas_list.html", {"minutas": minutas, "q": q})


@never_cache
def crear_minuta(request):
    ambientes = Ambiente.objects.all().order_by("num_ambiente")
    responsables = Instructor.objects.select_related("usuario_id_usuario").all()
    usuario, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied

    if request.method == "POST":
        try:
            ambiente_id = int(request.POST.get("ambiente_id"))
            responsable_id = int(request.POST.get("responsable_id"))
            fecha_recibo = datetime.strptime(request.POST.get("fecha_hora_recibo"), "%Y-%m-%dT%H:%M")
            fecha_entrega = datetime.strptime(request.POST.get("fecha_hora_entrega"), "%Y-%m-%dT%H:%M")
            estado = (request.POST.get("estado") or "").strip()
            novedad = (request.POST.get("novedad") or "").strip() or None
            descripcion = (request.POST.get("descripcion_min") or "").strip() or None
        except (TypeError, ValueError):
            messages.error(request, "Datos inválidos para crear la minuta.")
            return _render_guarda(
                request,
                "guarda/minuta_form.html",
                {"ambientes": ambientes, "responsables": responsables, "modo": "crear"},
            )

        guarda = GuardaSeguridad.objects.filter(usuario_id_usuario=usuario).first()
        if not guarda:
            messages.error(request, "Tu usuario no tiene registro activo como guarda.")
            return _render_guarda(request, "guarda/minuta_form.html", {"ambientes": ambientes, "responsables": responsables, "modo": "crear"})

        RegistroMinuta.objects.create(
            fecha_hora_recibo=fecha_recibo,
            fecha_hora_entrega=fecha_entrega,
            novedad=novedad,
            descripcion_min=descripcion,
            estado=estado,
            ambiente_id=ambiente_id,
            guarda_seguridad_usuario_id_usuario=guarda,
            responsable_id=responsable_id,
        )
        messages.success(request, "Minuta creada correctamente.")
        return _replace_redirect("guarda_minutas")

    return _render_guarda(
        request,
        "guarda/minuta_form.html",
        {"ambientes": ambientes, "responsables": responsables, "modo": "crear"},
    )


@never_cache
def editar_minuta(request, minuta_id):
    _, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    minuta = get_object_or_404(RegistroMinuta, pk=minuta_id)
    ambientes = Ambiente.objects.all().order_by("num_ambiente")
    responsables = Instructor.objects.select_related("usuario_id_usuario").all()
    if request.method == "POST":
        try:
            minuta.ambiente_id = int(request.POST.get("ambiente_id"))
            minuta.responsable_id = int(request.POST.get("responsable_id"))
            minuta.fecha_hora_recibo = datetime.strptime(request.POST.get("fecha_hora_recibo"), "%Y-%m-%dT%H:%M")
            minuta.fecha_hora_entrega = datetime.strptime(request.POST.get("fecha_hora_entrega"), "%Y-%m-%dT%H:%M")
            minuta.estado = (request.POST.get("estado") or "").strip()
            minuta.novedad = (request.POST.get("novedad") or "").strip() or None
            minuta.descripcion_min = (request.POST.get("descripcion_min") or "").strip() or None
            minuta.save()
            messages.success(request, "Minuta actualizada correctamente.")
            return _replace_redirect("guarda_minutas")
        except (TypeError, ValueError):
            messages.error(request, "No se pudo actualizar la minuta, revisa los campos.")

    return _render_guarda(
        request,
        "guarda/minuta_form.html",
        {"minuta": minuta, "ambientes": ambientes, "responsables": responsables, "modo": "editar"},
    )


@require_POST
@never_cache
def eliminar_minuta(request, minuta_id):
    _, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    minuta = get_object_or_404(RegistroMinuta, pk=minuta_id)
    minuta.delete()
    messages.success(request, "Minuta eliminada correctamente.")
    return _replace_redirect("guarda_minutas")


@never_cache
def listar_incidentes(request):
    q = _q_param(request)
    incidentes = (
        RegistroIncidente.objects.select_related("ambiente", "tipo_inc", "usuario_id_usuario")
        .all()
        .order_by("-fecha_incidente", "-hora_incidente")
    )
    incidentes = _filtrar_incidentes(incidentes, q)
    return _render_guarda(request, "guarda/incidentes_list.html", {"incidentes": incidentes, "q": q})


@never_cache
def crear_incidente(request):
    """El guarda solo consulta incidentes; el registro lo gestiona el instructor."""
    usuario, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    messages.info(
        request,
        "Los incidentes los registra el instructor desde su panel. Aquí solo puedes consultarlos y exportarlos.",
    )
    return _no_cache(redirect("guarda_incidentes"))


@never_cache
def editar_incidente(request, incidente_id):
    usuario, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    messages.info(
        request,
        "El guarda no puede modificar incidentes; solo consultarlos.",
    )
    return _no_cache(redirect("guarda_incidentes"))


@require_POST
@never_cache
def eliminar_incidente(request, incidente_id):
    usuario, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    messages.info(
        request,
        "El guarda no puede eliminar incidentes; solo consultarlos.",
    )
    return _no_cache(redirect("guarda_incidentes"))


@never_cache
def listar_traslados(request):
    q = _q_param(request)
    traslados = (
        TrasladoRecurso.objects.select_related("recurso", "ambiente_origen")
        .all()
        .order_by("-fecha_traslado")
    )
    traslados = _filtrar_traslados(traslados, q)
    return _render_guarda(request, "guarda/traslados_list.html", {"traslados": traslados, "q": q})


@never_cache
def crear_traslado(request):
    _, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    recursos = Recursos.objects.select_related("ambiente").all().order_by("nombre_recurso")
    ambientes = Ambiente.objects.all().order_by("num_ambiente")
    if request.method == "POST":
        try:
            fecha_traslado = datetime.strptime(request.POST.get("fecha_traslado"), "%Y-%m-%dT%H:%M")
            TrasladoRecurso.objects.create(
                recurso_id=int(request.POST.get("recurso_id")),
                ambiente_origen_id=int(request.POST.get("ambiente_origen_id")),
                ambiente_destino=int(request.POST.get("ambiente_destino")),
                fecha_traslado=fecha_traslado,
                observacion=(request.POST.get("observacion") or "").strip() or None,
            )
            messages.success(request, "Traslado creado correctamente.")
            return _replace_redirect("guarda_traslados")
        except (TypeError, ValueError, IntegrityError):
            messages.error(request, "No se pudo crear el traslado.")
    return _render_guarda(request, "guarda/traslado_form.html", {"recursos": recursos, "ambientes": ambientes, "modo": "crear"})


@never_cache
def editar_traslado(request, traslado_id):
    _, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    traslado = get_object_or_404(TrasladoRecurso, pk=traslado_id)
    recursos = Recursos.objects.select_related("ambiente").all().order_by("nombre_recurso")
    ambientes = Ambiente.objects.all().order_by("num_ambiente")
    if request.method == "POST":
        try:
            traslado.recurso_id = int(request.POST.get("recurso_id"))
            traslado.ambiente_origen_id = int(request.POST.get("ambiente_origen_id"))
            traslado.ambiente_destino = int(request.POST.get("ambiente_destino"))
            traslado.fecha_traslado = datetime.strptime(request.POST.get("fecha_traslado"), "%Y-%m-%dT%H:%M")
            traslado.observacion = (request.POST.get("observacion") or "").strip() or None
            traslado.save()
            messages.success(request, "Traslado actualizado correctamente.")
            return _replace_redirect("guarda_traslados")
        except (TypeError, ValueError, IntegrityError):
            messages.error(request, "No se pudo actualizar el traslado.")
    return _render_guarda(request, "guarda/traslado_form.html", {"traslado": traslado, "recursos": recursos, "ambientes": ambientes, "modo": "editar"})


@require_POST
@never_cache
def eliminar_traslado(request, traslado_id):
    _, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    traslado = get_object_or_404(TrasladoRecurso, pk=traslado_id)
    traslado.delete()
    messages.success(request, "Traslado eliminado correctamente.")
    return _replace_redirect("guarda_traslados")


@never_cache
def listar_ambientes(request):
    q = _q_param(request)
    ambientes = Ambiente.objects.all().order_by("id_ambiente")
    ambientes = _filtrar_ambientes(ambientes, q)
    return _render_guarda(request, "guarda/ambientes_list.html", {"ambientes": ambientes, "q": q})


@never_cache
def exportar_minutas_excel(request):
    q = _q_param(request)
    minutas = (
        RegistroMinuta.objects.select_related(
            "ambiente",
            "guarda_seguridad_usuario_id_usuario__usuario_id_usuario",
            "responsable__usuario_id_usuario",
        )
        .all()
        .order_by("-fecha_hora_recibo")
    )
    minutas = _filtrar_minutas(minutas, q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Minutas"

    headers = [
        "ID",
        "Ambiente",
        "Recibo",
        "Entrega",
        "Estado",
        "Novedad",
        "Descripción",
        "Responsable",
        "Guarda",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for m in minutas:
        responsable = getattr(m.responsable.usuario_id_usuario, "p_nombre", "") + " " + getattr(
            m.responsable.usuario_id_usuario, "p_apellido", ""
        )
        guarda = getattr(m.guarda_seguridad_usuario_id_usuario.usuario_id_usuario, "p_nombre", "") + " " + getattr(
            m.guarda_seguridad_usuario_id_usuario.usuario_id_usuario, "p_apellido", ""
        )
        ws.append(
            [
                m.id_minuta,
                m.ambiente.num_ambiente if m.ambiente_id else "",
                m.fecha_hora_recibo.strftime("%Y-%m-%d %H:%M:%S") if m.fecha_hora_recibo else "",
                m.fecha_hora_entrega.strftime("%Y-%m-%d %H:%M:%S") if m.fecha_hora_entrega else "",
                m.estado or "",
                m.novedad or "",
                m.descripcion_min or "",
                responsable.strip(),
                guarda.strip(),
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="minutas_{_safe_q_part(q)}.xlsx"'
    return _no_cache(resp)


@never_cache
def exportar_minutas_pdf(request):
    q = _q_param(request)
    minutas = (
        RegistroMinuta.objects.select_related(
            "ambiente",
            "guarda_seguridad_usuario_id_usuario__usuario_id_usuario",
            "responsable__usuario_id_usuario",
        )
        .all()
        .order_by("-fecha_hora_recibo")
    )
    minutas = _filtrar_minutas(minutas, q)

    headers = ["ID", "Ambiente", "Recibo", "Entrega", "Estado", "Novedad", "Descripción"]
    rows = []
    for m in minutas:
        rows.append(
            [
                str(m.id_minuta),
                str(m.ambiente.num_ambiente) if m.ambiente_id else "",
                m.fecha_hora_recibo.strftime("%Y-%m-%d %H:%M") if m.fecha_hora_recibo else "",
                m.fecha_hora_entrega.strftime("%Y-%m-%d %H:%M") if m.fecha_hora_entrega else "",
                m.estado or "",
                m.novedad or "",
                m.descripcion_min or "",
            ]
        )
    resp = landscape_pdf_response(
        "Reporte de Minutas",
        q,
        headers,
        rows,
        [0.05, 0.07, 0.11, 0.11, 0.09, 0.22, 0.35],
        f"minutas_{_safe_q_part(q)}.pdf",
    )
    return _no_cache(resp)


@never_cache
def exportar_incidentes_excel(request):
    q = _q_param(request)
    incidentes = (
        RegistroIncidente.objects.select_related("ambiente", "tipo_inc", "usuario_id_usuario")
        .all()
        .order_by("-fecha_incidente", "-hora_incidente")
    )
    incidentes = _filtrar_incidentes(incidentes, q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Incidentes"

    headers = ["ID", "Fecha", "Hora", "Ambiente", "Tipo", "Descripción", "Usuario"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i in incidentes:
        usuario = f"{i.usuario_id_usuario.p_nombre} {i.usuario_id_usuario.p_apellido}"
        ws.append(
            [
                i.id_incidente,
                i.fecha_incidente.isoformat() if i.fecha_incidente else "",
                str(i.hora_incidente) if i.hora_incidente else "",
                i.ambiente.num_ambiente if i.ambiente_id else "",
                i.tipo_inc.tipo_incidente if i.tipo_inc_id else "",
                i.descripcion or "",
                usuario.strip(),
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="incidentes_{_safe_q_part(q)}.xlsx"'
    return _no_cache(resp)


@never_cache
def exportar_incidentes_pdf(request):
    q = _q_param(request)
    incidentes = (
        RegistroIncidente.objects.select_related("ambiente", "tipo_inc", "usuario_id_usuario")
        .all()
        .order_by("-fecha_incidente", "-hora_incidente")
    )
    incidentes = _filtrar_incidentes(incidentes, q)

    headers = ["ID", "Fecha", "Hora", "Ambiente", "Tipo", "Descripción", "Usuario"]
    rows = []
    for i in incidentes:
        usuario = f"{i.usuario_id_usuario.p_nombre} {i.usuario_id_usuario.p_apellido}".strip()
        rows.append(
            [
                str(i.id_incidente),
                i.fecha_incidente.isoformat() if i.fecha_incidente else "",
                str(i.hora_incidente) if i.hora_incidente else "",
                str(i.ambiente.num_ambiente) if i.ambiente_id else "",
                i.tipo_inc.tipo_incidente if i.tipo_inc_id else "",
                i.descripcion or "",
                usuario,
            ]
        )
    resp = landscape_pdf_response(
        "Reporte de Incidentes",
        q,
        headers,
        rows,
        [0.06, 0.09, 0.07, 0.08, 0.10, 0.35, 0.25],
        f"incidentes_{_safe_q_part(q)}.pdf",
    )
    return _no_cache(resp)


@never_cache
def exportar_traslados_excel(request):
    q = _q_param(request)
    traslados = (
        TrasladoRecurso.objects.select_related("recurso", "ambiente_origen")
        .all()
        .order_by("-fecha_traslado")
    )
    traslados = _filtrar_traslados(traslados, q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Traslados"

    headers = ["ID", "Recurso", "Serial", "Origen", "Destino", "Fecha", "Observación"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for t in traslados:
        ws.append(
            [
                t.id_traslado,
                t.recurso.nombre_recurso if t.recurso_id else "",
                t.recurso.serial_recurso if t.recurso_id else "",
                t.ambiente_origen.num_ambiente if t.ambiente_origen_id else "",
                t.ambiente_destino or "",
                t.fecha_traslado.strftime("%Y-%m-%d %H:%M:%S") if t.fecha_traslado else "",
                t.observacion or "",
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="traslados_{_safe_q_part(q)}.xlsx"'
    return _no_cache(resp)


@never_cache
def exportar_traslados_pdf(request):
    q = _q_param(request)
    traslados = (
        TrasladoRecurso.objects.select_related("recurso", "ambiente_origen")
        .all()
        .order_by("-fecha_traslado")
    )
    traslados = _filtrar_traslados(traslados, q)

    headers = ["ID", "Recurso", "Serial", "Origen", "Destino", "Fecha", "Observación"]
    rows = []
    for t in traslados:
        rows.append(
            [
                str(t.id_traslado),
                t.recurso.nombre_recurso if t.recurso_id else "",
                t.recurso.serial_recurso if t.recurso_id else "",
                str(t.ambiente_origen.num_ambiente) if t.ambiente_origen_id else "",
                str(t.ambiente_destino or ""),
                t.fecha_traslado.strftime("%Y-%m-%d %H:%M:%S") if t.fecha_traslado else "",
                t.observacion or "",
            ]
        )
    resp = landscape_pdf_response(
        "Reporte de Traslados",
        q,
        headers,
        rows,
        [0.07, 0.18, 0.14, 0.10, 0.10, 0.14, 0.27],
        f"traslados_{_safe_q_part(q)}.pdf",
    )
    return _no_cache(resp)


@never_cache
def exportar_ambientes_excel(request):
    q = _q_param(request)
    ambientes = _filtrar_ambientes(Ambiente.objects.all(), q).order_by("id_ambiente")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ambientes"

    headers = ["ID", "Número", "Capacidad", "Tipo", "Estado"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for a in ambientes:
        ws.append([a.id_ambiente, a.num_ambiente, a.capacidad, a.tipo_ambiente or "", a.estado or ""])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="ambientes_{_safe_q_part(q)}.xlsx"'
    return _no_cache(resp)


@never_cache
def exportar_ambientes_pdf(request):
    q = _q_param(request)
    ambientes = _filtrar_ambientes(Ambiente.objects.all(), q).order_by("id_ambiente")

    headers = ["ID", "Número", "Capacidad", "Tipo", "Estado"]
    rows = []
    for a in ambientes:
        rows.append(
            [
                str(a.id_ambiente),
                str(a.num_ambiente),
                str(a.capacidad),
                a.tipo_ambiente or "",
                a.estado or "",
            ]
        )
    resp = landscape_pdf_response(
        "Reporte de Ambientes",
        q,
        headers,
        rows,
        [0.12, 0.14, 0.14, 0.35, 0.25],
        f"ambientes_{_safe_q_part(q)}.pdf",
    )
    return _no_cache(resp)


@never_cache
def crear_ambiente(request):
    """El guarda solo consulta el estado de los ambientes."""
    usuario, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    messages.info(
        request,
        "El guarda solo puede consultar ambientes. La creación y el mantenimiento los gestiona personal autorizado (por ejemplo administración).",
    )
    return _no_cache(redirect("guarda_ambientes"))


@never_cache
def editar_ambiente(request, ambiente_id):
    usuario, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    messages.info(
        request,
        "El guarda no puede editar ambientes; solo consultarlos.",
    )
    return _no_cache(redirect("guarda_ambientes"))


@require_POST
@never_cache
def eliminar_ambiente(request, ambiente_id):
    usuario, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied
    messages.info(
        request,
        "El guarda no puede eliminar ambientes; solo consultarlos.",
    )
    return _no_cache(redirect("guarda_ambientes"))


@never_cache
def mi_perfil(request):
    usuario, denied = _acceso_guarda_o_login(request)
    if denied:
        return denied

    user_rol = UserRol.objects.filter(id_usuario=usuario).select_related("id_rol").first()
    rol_nombre = (user_rol.id_rol.nombre_rol or "").strip() if user_rol else "Guarda de Seguridad"

    if request.method == "POST":
        usuario.p_nombre = (request.POST.get("p_nombre") or "").strip()
        usuario.s_nombre = (request.POST.get("s_nombre") or "").strip() or None
        usuario.p_apellido = (request.POST.get("p_apellido") or "").strip()
        usuario.s_apellido = (request.POST.get("s_apellido") or "").strip() or None
        usuario.correo = (request.POST.get("correo") or "").strip()
        try:
            with transaction.atomic():
                usuario.save()
            messages.success(request, "Perfil actualizado correctamente.")
            return _replace_redirect("guarda_perfil")
        except IntegrityError:
            messages.error(request, "No se pudo actualizar el perfil.")

    return _render_guarda(
        request,
        "guarda/perfil.html",
        {"perfil": usuario, "rol_nombre": rol_nombre},
    )
