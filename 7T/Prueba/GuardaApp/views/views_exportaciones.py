from datetime import datetime
from io import BytesIO
import json
import re

from django.contrib import messages
from django.db import IntegrityError, transaction
from django.db.models import Max, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from LoginApp.models import (
    Ambiente,
    GuardaSeguridad,
    Instructor,
    Recursos,
    RegistroIncidente,
    RegistroMinuta,
    Rol,
    TipoIncidente,
    TrasladoRecurso,
    UserRol,
    Usuario,
)



from .utils import *
from LoginApp.decorators import login_requerido, rol_requerido

@never_cache
def exportar_minutas_excel(request):
    q = _q_param(request)
    minutas = (
        RegistroMinuta.objects.select_related(
            "ambiente",
            "guarda_seguridad_usuario_id_usuario__usuario_id_usuario",
            "responsable__usuario_id_usuario",
        )
        .all()
        .order_by("-fecha_hora_recibo")
    )
    minutas = _filtrar_minutas(minutas, q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Minutas"

    headers = [
        "ID",
        "Ambiente",
        "Recibo",
        "Entrega",
        "Estado",
        "Novedad",
        "Descripción",
        "Responsable",
        "Guarda",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for m in minutas:
        responsable = getattr(m.responsable.usuario_id_usuario, "p_nombre", "") + " " + getattr(
            m.responsable.usuario_id_usuario, "p_apellido", ""
        )
        guarda = getattr(m.guarda_seguridad_usuario_id_usuario.usuario_id_usuario, "p_nombre", "") + " " + getattr(
            m.guarda_seguridad_usuario_id_usuario.usuario_id_usuario, "p_apellido", ""
        )
        ws.append(
            [
                m.id_minuta,
                m.ambiente.num_ambiente if m.ambiente_id else "",
                m.fecha_hora_recibo.strftime("%Y-%m-%d %H:%M:%S") if m.fecha_hora_recibo else "",
                m.fecha_hora_entrega.strftime("%Y-%m-%d %H:%M:%S") if m.fecha_hora_entrega else "",
                m.estado or "",
                m.novedad or "",
                m.descripcion_min or "",
                responsable.strip(),
                guarda.strip(),
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="minutas_{_safe_q_part(q)}.xlsx"'
    return _no_cache(resp)


@never_cache
def exportar_minutas_pdf(request):
    q = _q_param(request)
    minutas = (
        RegistroMinuta.objects.select_related(
            "ambiente",
            "guarda_seguridad_usuario_id_usuario__usuario_id_usuario",
            "responsable__usuario_id_usuario",
        )
        .all()
        .order_by("-fecha_hora_recibo")
    )
    minutas = _filtrar_minutas(minutas, q)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)

    elements = [Paragraph("Reporte de Minutas", styles["Title"]), Spacer(1, 12)]
    subtitle = f"Filtro: {q}" if q else "Sin filtro"
    elements.append(Paragraph(subtitle, styles["BodyText"]))
    elements.append(Spacer(1, 12))

    data = [
        [
            "ID",
            "Ambiente",
            "Recibo",
            "Entrega",
            "Estado",
            "Novedad",
            "Descripción",
        ]
    ]
    for m in minutas:
        data.append(
            [
                str(m.id_minuta),
                str(m.ambiente.num_ambiente) if m.ambiente_id else "",
                m.fecha_hora_recibo.strftime("%Y-%m-%d %H:%M:%S") if m.fecha_hora_recibo else "",
                m.fecha_hora_entrega.strftime("%Y-%m-%d %H:%M:%S") if m.fecha_hora_entrega else "",
                m.estado or "",
                m.novedad or "",
                (m.descripcion_min or "")[:300],
            ]
        )

    table = Table(data, repeatRows=1, colWidths=[1.0 * cm, 2.2 * cm, 2.6 * cm, 2.6 * cm, 2.1 * cm, 2.2 * cm, 4.0 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)

    pdf_bytes = buf.getvalue()
    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="minutas_{_safe_q_part(q)}.pdf"'
    return _no_cache(resp)


@never_cache
def exportar_incidentes_excel(request):
    q = _q_param(request)
    incidentes = (
        RegistroIncidente.objects.select_related("ambiente", "tipo_inc", "usuario_id_usuario")
        .all()
        .order_by("-fecha_incidente", "-hora_incidente")
    )
    incidentes = _filtrar_incidentes(incidentes, q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Incidentes"

    headers = ["ID", "Fecha", "Hora", "Ambiente", "Tipo", "Descripción", "Usuario"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i in incidentes:
        usuario = f"{i.usuario_id_usuario.p_nombre} {i.usuario_id_usuario.p_apellido}"
        ws.append(
            [
                i.id_incidente,
                i.fecha_incidente.isoformat() if i.fecha_incidente else "",
                str(i.hora_incidente) if i.hora_incidente else "",
                i.ambiente.num_ambiente if i.ambiente_id else "",
                i.tipo_inc.tipo_incidente if i.tipo_inc_id else "",
                i.descripcion or "",
                usuario.strip(),
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="incidentes_{_safe_q_part(q)}.xlsx"'
    return _no_cache(resp)


@never_cache
def exportar_incidentes_pdf(request):
    q = _q_param(request)
    incidentes = (
        RegistroIncidente.objects.select_related("ambiente", "tipo_inc", "usuario_id_usuario")
        .all()
        .order_by("-fecha_incidente", "-hora_incidente")
    )
    incidentes = _filtrar_incidentes(incidentes, q)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)

    elements = [Paragraph("Reporte de Incidentes", styles["Title"]), Spacer(1, 12)]
    subtitle = f"Filtro: {q}" if q else "Sin filtro"
    elements.append(Paragraph(subtitle, styles["BodyText"]))
    elements.append(Spacer(1, 12))

    data = [["ID", "Fecha", "Hora", "Ambiente", "Tipo", "Descripción", "Usuario"]]
    for i in incidentes:
        usuario = f"{i.usuario_id_usuario.p_nombre} {i.usuario_id_usuario.p_apellido}".strip()
        data.append(
            [
                str(i.id_incidente),
                i.fecha_incidente.isoformat() if i.fecha_incidente else "",
                str(i.hora_incidente) if i.hora_incidente else "",
                str(i.ambiente.num_ambiente) if i.ambiente_id else "",
                i.tipo_inc.tipo_incidente if i.tipo_inc_id else "",
                (i.descripcion or "")[:300],
                usuario,
            ]
        )

    table = Table(
        data,
        repeatRows=1,
        colWidths=[1.0 * cm, 2.0 * cm, 1.6 * cm, 2.2 * cm, 2.2 * cm, 4.2 * cm, 2.6 * cm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)

    pdf_bytes = buf.getvalue()
    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="incidentes_{_safe_q_part(q)}.pdf"'
    return _no_cache(resp)


@never_cache
def exportar_traslados_excel(request):
    q = _q_param(request)
    traslados = (
        TrasladoRecurso.objects.select_related("recurso", "ambiente_origen")
        .all()
        .order_by("-fecha_traslado")
    )
    traslados = _filtrar_traslados(traslados, q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Traslados"

    headers = ["ID", "Recurso", "Serial", "Origen", "Destino", "Fecha", "Observación"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for t in traslados:
        ws.append(
            [
                t.id_traslado,
                t.recurso.nombre_recurso if t.recurso_id else "",
                t.recurso.serial_recurso if t.recurso_id else "",
                t.ambiente_origen.num_ambiente if t.ambiente_origen_id else "",
                t.ambiente_destino or "",
                t.fecha_traslado.strftime("%Y-%m-%d %H:%M:%S") if t.fecha_traslado else "",
                t.observacion or "",
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="traslados_{_safe_q_part(q)}.xlsx"'
    return _no_cache(resp)


@never_cache
def exportar_traslados_pdf(request):
    q = _q_param(request)
    traslados = (
        TrasladoRecurso.objects.select_related("recurso", "ambiente_origen")
        .all()
        .order_by("-fecha_traslado")
    )
    traslados = _filtrar_traslados(traslados, q)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)

    elements = [Paragraph("Reporte de Traslados", styles["Title"]), Spacer(1, 12)]
    subtitle = f"Filtro: {q}" if q else "Sin filtro"
    elements.append(Paragraph(subtitle, styles["BodyText"]))
    elements.append(Spacer(1, 12))

    data = [["ID", "Recurso", "Serial", "Origen", "Destino", "Fecha", "Observación"]]
    for t in traslados:
        data.append(
            [
                str(t.id_traslado),
                t.recurso.nombre_recurso if t.recurso_id else "",
                t.recurso.serial_recurso if t.recurso_id else "",
                str(t.ambiente_origen.num_ambiente) if t.ambiente_origen_id else "",
                str(t.ambiente_destino or ""),
                t.fecha_traslado.strftime("%Y-%m-%d %H:%M:%S") if t.fecha_traslado else "",
                (t.observacion or "")[:300],
            ]
        )

    table = Table(
        data,
        repeatRows=1,
        colWidths=[1.0 * cm, 3.0 * cm, 3.0 * cm, 2.0 * cm, 1.8 * cm, 2.6 * cm, 3.8 * cm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)

    pdf_bytes = buf.getvalue()
    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="traslados_{_safe_q_part(q)}.pdf"'
    return _no_cache(resp)


@never_cache
def exportar_ambientes_excel(request):
    q = _q_param(request)
    ambientes = _filtrar_ambientes(Ambiente.objects.all(), q).order_by("id_ambiente")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ambientes"

    headers = ["ID", "Número", "Capacidad", "Tipo", "Estado"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for a in ambientes:
        ws.append([a.id_ambiente, a.num_ambiente, a.capacidad, a.tipo_ambiente or "", a.estado or ""])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="ambientes_{_safe_q_part(q)}.xlsx"'
    return _no_cache(resp)


@never_cache
def exportar_ambientes_pdf(request):
    q = _q_param(request)
    ambientes = _filtrar_ambientes(Ambiente.objects.all(), q).order_by("id_ambiente")

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)

    elements = [Paragraph("Reporte de Ambientes", styles["Title"]), Spacer(1, 12)]
    subtitle = f"Filtro: {q}" if q else "Sin filtro"
    elements.append(Paragraph(subtitle, styles["BodyText"]))
    elements.append(Spacer(1, 12))

    data = [["ID", "Número", "Capacidad", "Tipo", "Estado"]]
    for a in ambientes:
        data.append([str(a.id_ambiente), str(a.num_ambiente), str(a.capacidad), a.tipo_ambiente or "", a.estado or ""])

    table = Table(data, repeatRows=1, colWidths=[1.4 * cm, 2.2 * cm, 2.0 * cm, 3.8 * cm, 2.5 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)

    pdf_bytes = buf.getvalue()
    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="ambientes_{_safe_q_part(q)}.pdf"'
    return _no_cache(resp)


