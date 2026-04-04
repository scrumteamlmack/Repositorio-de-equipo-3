import hashlib
from collections import defaultdict
from io import BytesIO
from urllib.parse import urlencode

from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth import logout
from django.db import IntegrityError, transaction
from django.db.models import Max, Q
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from Prueba.report_utils import landscape_pdf_response, maybe_int, q_param as admin_q, safe_q_part as admin_safe_q

from LoginApp.models import (
    Usuario,
    UserRol,
    Ficha,
    Rol,
    Programas,
    Recursos,
    Ambiente,
    Instructor,
    Jornada,
    Modalidad,
    Coordinacion,
    TipoRecurso,
)

MAX_MEDIUMINT = 8_388_607

def _hash_nueva_contrasena(plain: str) -> str:
    """Mismo criterio que registros recientes en BD: SHA-256 en hexadecimal."""
    return hashlib.sha256(plain.encode("utf-8")).hexdigest()


def _usuario_a_formulario(u: Usuario, rol_id=None):
    if rol_id is None:
        ur = UserRol.objects.filter(id_usuario=u).first()
        rol_id = ur.id_rol_id if ur else ""
    return {
        "PNombre": u.p_nombre,
        "SNombre": u.s_nombre or "",
        "PApellido": u.p_apellido,
        "SApellido": u.s_apellido or "",
        "tipoDocumento": u.tipo_documento,
        "numDocumento": u.num_documento,
        "correo": u.correo,
        "rolSeleccionado": str(rol_id) if rol_id != "" else "",
    }


EMPTY_USER = {
    "PNombre": "",
    "SNombre": "",
    "PApellido": "",
    "SApellido": "",
    "tipoDocumento": "",
    "numDocumento": "",
    "correo": "",
    "rolSeleccionado": "",
}

EMPTY_RESOURCE = {
    "nombre": "",
    "serial": "",
    "numero": "",
    "idTipoRecurso": "",
    "idAmbiente": "",
    "estado": "",
    "observacion": "",
}

EMPTY_AMBIENTE = {
    "numero": "",
    "capacidad": "",
    "tipo": "",
    "estado": "",
}

EMPTY_PROGRAMA = {
    "nombrePrograma": "",
    "nivelFormacion": "",
    "duracion": "",
    "jornadaId": "",
    "modalidadId": "",
    "coordinacionId": "",
}

EMPTY_FICHA = {
    "numFicha": "",
}


def _render_admin(request, template_name, extra_context=None):
    context = {
        "roles": [],
        "usuarios": [],
        "fichas": [],
        "programas": [],
        "recursos": [],
        "ambientes": [],
        "tipos": [],
        "jornadas": [],
        "modalidades": [],
        "coordinaciones": [],
        "instructores": [],
        "turnos": ["Manana", "Tarde", "Noche"],
        "usuario": EMPTY_USER.copy(),
        "recurso": EMPTY_RESOURCE.copy(),
        "ambiente": EMPTY_AMBIENTE.copy(),
        "programa": EMPTY_PROGRAMA.copy(),
        "ficha": EMPTY_FICHA.copy(),
    }
    if extra_context:
        context.update(extra_context)
    return render(request, template_name, context)


def cerrar_sesion(request):
    # Este proyecto maneja sesión propia con `request.session['usuario_id']`.
    # Siempre la limpiamos, incluso si no se usa `django.contrib.auth`.
    request.session.pop("usuario_id", None)
    try:
        request.session.flush()
    except Exception:
        # Si el backend de sesión no permite flush por algún motivo,
        # al menos ya eliminamos la clave principal.
        pass
    if request.user.is_authenticated:
        logout(request)
    messages.success(request, "Sesión cerrada correctamente.")
    return redirect("siza")


def admin_index(request):
    return _render_admin(request, "admin_index.html")


def uadmin(request):
    return redirect("admin_index")


def index_admin(request):
    return _render_admin(request, "indexAdmin.html")


def siza(request):
    return render(request, "siza.html")


def perfil(request):
    uid = request.session.get("usuario_id")
    if not uid:
        messages.warning(request, "Inicie sesión para ver su perfil.")
        login_url = reverse("login")
        return redirect(f"{login_url}?{urlencode({'next': reverse('perfil')})}")

    u = get_object_or_404(Usuario, pk=uid)
    roles_qs = UserRol.objects.filter(id_usuario=u).select_related("id_rol")
    roles_detalle = ", ".join(ur.id_rol.nombre_rol for ur in roles_qs) or "Sin rol asignado"
    es_admin = any(
        ur.id_rol.nombre_rol.lower() in ("admin", "administrador") for ur in roles_qs
    )
    nombre_completo = " ".join(
        filter(None, [u.p_nombre, u.s_nombre, u.p_apellido, u.s_apellido])
    ).strip() or u.correo

    return render(
        request,
        "perfil.html",
        {
            "usuario_perfil": {
                "id": u.id_usuario,
                "nombre_completo": nombre_completo,
                "p_nombre": u.p_nombre,
                "s_nombre": u.s_nombre,
                "p_apellido": u.p_apellido,
                "s_apellido": u.s_apellido,
                "tipo_documento": u.tipo_documento,
                "num_documento": u.num_documento,
                "correo": u.correo,
                "roles": roles_detalle,
            },
            "puede_editar_perfil": es_admin,
        },
    )


def listar_usuarios(request):
    q = admin_q(request)
    roles_por_usuario = defaultdict(list)
    for ur in UserRol.objects.select_related("id_rol").all():
        roles_por_usuario[ur.id_usuario_id].append(ur.id_rol.nombre_rol)

    qs = Usuario.objects.all().order_by("id_usuario")
    if q:
        nu = maybe_int(q)
        cond = (
            Q(p_nombre__icontains=q)
            | Q(s_nombre__icontains=q)
            | Q(p_apellido__icontains=q)
            | Q(s_apellido__icontains=q)
            | Q(correo__icontains=q)
            | Q(tipo_documento__icontains=q)
            | Q(userrol_set__id_rol__nombre_rol__icontains=q)
        )
        if nu is not None:
            cond |= Q(id_usuario=nu) | Q(num_documento=nu)
        qs = qs.filter(cond).distinct()

    filas = []
    for u in qs:
        nombres = " ".join(filter(None, [u.p_nombre, u.s_nombre])).strip()
        apellidos = " ".join(filter(None, [u.p_apellido, u.s_apellido])).strip()
        roles = roles_por_usuario.get(u.id_usuario, [])
        filas.append(
            {
                "id": u.id_usuario,
                "first_name": nombres or "—",
                "last_name": apellidos or "—",
                "email": u.correo,
                "rol": ", ".join(roles) if roles else "—",
            }
        )
    return _render_admin(request, "listarUsuarios.html", {"usuarios": filas, "q": q})


def form_usuario(request):
    roles = list(Rol.objects.all().order_by("id_rol"))
    if request.method == "POST":
        p_nombre = request.POST.get("PNombre", "").strip()
        s_nombre = request.POST.get("SNombre", "").strip()
        p_apellido = request.POST.get("PApellido", "").strip()
        s_apellido = request.POST.get("SApellido", "").strip()
        tipo_doc = request.POST.get("tipoDocumento", "").strip()
        correo = request.POST.get("correo", "").strip()
        password = (request.POST.get("pass") or "").strip()
        rol_raw = request.POST.get("rolSeleccionado")
        num_raw = request.POST.get("numDocumento")

        try:
            num_doc = int(num_raw)
        except (TypeError, ValueError):
            messages.error(request, "Número de documento inválido.")
            return _render_admin(
                request,
                "formUsuario.html",
                {
                    "roles": roles,
                    "usuario": {
                        "PNombre": p_nombre,
                        "SNombre": s_nombre,
                        "PApellido": p_apellido,
                        "SApellido": s_apellido,
                        "tipoDocumento": tipo_doc,
                        "numDocumento": num_raw or "",
                        "correo": correo,
                        "rolSeleccionado": str(rol_raw) if rol_raw else "",
                    },
                },
            )

        if not p_nombre or not p_apellido or not tipo_doc:
            messages.error(request, "Complete los campos obligatorios.")
            return _render_admin(request, "formUsuario.html", {"roles": roles})
        if not password:
            messages.error(request, "La contraseña es obligatoria al crear usuario.")
            return _render_admin(request, "formUsuario.html", {"roles": roles})
        if not rol_raw:
            messages.error(request, "Seleccione un rol.")
            return _render_admin(request, "formUsuario.html", {"roles": roles})
        try:
            rol_id = int(rol_raw)
        except ValueError:
            messages.error(request, "Rol no válido.")
            return _render_admin(request, "formUsuario.html", {"roles": roles})
        if not Rol.objects.filter(pk=rol_id).exists():
            messages.error(request, "El rol indicado no existe.")
            return _render_admin(request, "formUsuario.html", {"roles": roles})
        if Usuario.objects.filter(num_documento=num_doc).exists():
            messages.error(request, "Ya existe un usuario con ese número de documento.")
            return _render_admin(request, "formUsuario.html", {"roles": roles})

        try:
            with transaction.atomic():
                u = Usuario.objects.create(
                    p_nombre=p_nombre,
                    s_nombre=s_nombre or None,
                    p_apellido=p_apellido,
                    s_apellido=s_apellido or None,
                    tipo_documento=tipo_doc,
                    num_documento=num_doc,
                    correo=correo,
                    contrasena=_hash_nueva_contrasena(password),
                )
                UserRol.objects.create(id_usuario=u, id_rol_id=rol_id)
        except IntegrityError: 
            messages.error(
                request,
                "No se pudo crear el usuario. Revise que el documento o correo no estén duplicados.",
            )
            return _render_admin(request, "formUsuario.html", {"roles": roles})

        messages.success(request, "Usuario registrado correctamente.")
        return redirect("listar_usuarios")

    return _render_admin(request, "formUsuario.html", {"roles": roles})


def crear_usuario(request):
    return form_usuario(request)


def editar_usuario(request, usuario_id):
    u = get_object_or_404(Usuario, pk=usuario_id)
    roles = list(Rol.objects.all().order_by("id_rol"))
    ur = UserRol.objects.filter(id_usuario=u).first()
    rol_actual = ur.id_rol_id if ur else ""

    if request.method == "POST":
        p_nombre = request.POST.get("PNombre", "").strip()
        s_nombre = request.POST.get("SNombre", "").strip()
        p_apellido = request.POST.get("PApellido", "").strip()
        s_apellido = request.POST.get("SApellido", "").strip()
        tipo_doc = request.POST.get("tipoDocumento", "").strip()
        correo = request.POST.get("correo", "").strip()
        password = (request.POST.get("pass") or "").strip()
        rol_raw = request.POST.get("rolSeleccionado")

        if not p_nombre or not p_apellido or not tipo_doc or not correo:
            messages.error(request, "Complete los campos obligatorios.")
            return _render_admin(
                request,
                "editarUsuario.html",
                {
                    "roles": roles,
                    "usuario": _usuario_a_formulario(u, rol_actual),
                    "idUsuarioEditar": usuario_id,
                },
            )
        if not rol_raw:
            messages.error(request, "Seleccione un rol.")
            return _render_admin(
                request,
                "editarUsuario.html",
                {
                    "roles": roles,
                    "usuario": _usuario_a_formulario(u, rol_actual),
                    "idUsuarioEditar": usuario_id,
                },
            )
        try:
            rol_id = int(rol_raw)
        except ValueError:
            messages.error(request, "Rol no válido.")
            return _render_admin(
                request,
                "editarUsuario.html",
                {
                    "roles": roles,
                    "usuario": _usuario_a_formulario(u, rol_actual),
                    "idUsuarioEditar": usuario_id,
                },
            )
        if not Rol.objects.filter(pk=rol_id).exists():
            messages.error(request, "El rol indicado no existe.")
            return _render_admin(
                request,
                "editarUsuario.html",
                {
                    "roles": roles,
                    "usuario": _usuario_a_formulario(u, rol_actual),
                    "idUsuarioEditar": usuario_id,
                },
            )

        u.p_nombre = p_nombre
        u.s_nombre = s_nombre or None
        u.p_apellido = p_apellido
        u.s_apellido = s_apellido or None
        u.tipo_documento = tipo_doc
        u.correo = correo
        if password:
            u.contrasena = _hash_nueva_contrasena(password)

        try:
            with transaction.atomic():
                u.save()
                UserRol.objects.filter(id_usuario=u).delete()
                UserRol.objects.create(id_usuario=u, id_rol_id=rol_id)
        except IntegrityError:
            messages.error(request, "No se pudo actualizar el usuario.")
            return _render_admin(
                request,
                "editarUsuario.html",
                {
                    "roles": roles,
                    "usuario": _usuario_a_formulario(u, rol_actual),
                    "idUsuarioEditar": usuario_id,
                },
            )

        messages.success(request, "Usuario actualizado correctamente.")
        return redirect("listar_usuarios")

    return _render_admin(
        request,
        "editarUsuario.html",
        {
            "roles": roles,
            "usuario": _usuario_a_formulario(u, rol_actual),
            "idUsuarioEditar": usuario_id,
        },
    )


@require_POST
def eliminar_usuario(request, usuario_id):
    u = Usuario.objects.filter(pk=usuario_id).first()
    if not u:
        messages.error(request, "Usuario no encontrado.")
        return redirect("listar_usuarios")
    try:
        with transaction.atomic():
            UserRol.objects.filter(id_usuario=u).delete()
            u.delete()
    except IntegrityError:
        messages.error(
            request,
            "No se puede eliminar: el usuario tiene datos enlazados (aprendiz, instructor, incidentes, etc.).",
        )
        return redirect("listar_usuarios")
    messages.success(request, "Usuario eliminado correctamente.")
    return redirect("listar_usuarios")


def form_instructor(request):
    return _render_admin(request, "formInstructor.html")


def form_coordinador(request):
    return _render_admin(request, "formCoordinador.html")


def form_guarda(request):
    return _render_admin(request, "formGuarda.html")


def _instructores_para_select():
    filas = []
    qs = Instructor.objects.select_related("usuario_id_usuario").order_by(
        "usuario_id_usuario__p_apellido",
        "usuario_id_usuario__p_nombre",
    )
    for inst in qs:
        u = inst.usuario_id_usuario
        nombre = " ".join(filter(None, [u.p_nombre, u.p_apellido])).strip() or f"ID {inst.pk}"
        filas.append({"id": inst.pk, "nombre": nombre})
    return filas


def listar_fichas(request):
    q = admin_q(request)
    qs = Ficha.objects.select_related(
        "instructor_usuario_id_usuario__usuario_id_usuario"
    ).order_by("num_ficha")
    if q:
        nu = maybe_int(q)
        uq = (
            Q(instructor_usuario_id_usuario__usuario_id_usuario__p_nombre__icontains=q)
            | Q(instructor_usuario_id_usuario__usuario_id_usuario__s_nombre__icontains=q)
            | Q(instructor_usuario_id_usuario__usuario_id_usuario__p_apellido__icontains=q)
            | Q(instructor_usuario_id_usuario__usuario_id_usuario__s_apellido__icontains=q)
        )
        if nu is not None:
            uq |= Q(idficha=nu) | Q(num_ficha=nu)
        qs = qs.filter(uq).distinct()

    filas = []
    for f in qs:
        u = f.instructor_usuario_id_usuario.usuario_id_usuario
        instructor = " ".join(filter(None, [u.p_nombre, u.p_apellido])).strip()
        filas.append(
            {
                "id": f.idficha,
                "num_ficha": f.num_ficha,
                "instructor": instructor or "—",
            }
        )
    return _render_admin(request, "listarFichas.html", {"fichas": filas, "q": q})


def crear_ficha(request):
    instructores = _instructores_para_select()
    if request.method == "POST":
        num_raw = (request.POST.get("numFicha") or "").strip()
        inst_raw = request.POST.get("instructorId")
        try:
            num_ficha = int(num_raw)
        except ValueError:
            messages.error(request, "Número de ficha inválido.")
            return _render_admin(
                request,
                "formFicha.html",
                {"instructores": instructores, "ficha": {"numFicha": num_raw, "instructorId": inst_raw or ""}},
            )
        if num_ficha < 1 or num_ficha > MAX_MEDIUMINT:
            messages.error(
                request,
                f"El número de ficha debe estar entre 1 y {MAX_MEDIUMINT}.",
            )
            return _render_admin(
                request,
                "formFicha.html",
                {"instructores": instructores, "ficha": {"numFicha": num_raw, "instructorId": inst_raw or ""}},
            )
        if not inst_raw:
            messages.error(request, "Seleccione un instructor.")
            return _render_admin(
                request,
                "formFicha.html",
                {"instructores": instructores, "ficha": {"numFicha": num_raw, "instructorId": ""}},
            )
        try:
            inst_id = int(inst_raw)
        except ValueError:
            messages.error(request, "Instructor no válido.")
            return _render_admin(
                request,
                "formFicha.html",
                {"instructores": instructores, "ficha": {"numFicha": num_raw, "instructorId": inst_raw}},
            )
        if not Instructor.objects.filter(pk=inst_id).exists():
            messages.error(request, "El instructor elegido no existe.")
            return _render_admin(
                request,
                "formFicha.html",
                {"instructores": instructores, "ficha": {"numFicha": num_raw, "instructorId": str(inst_id)}},
            )

        siguiente = (Ficha.objects.aggregate(m=Max("idficha"))["m"] or 0) + 1
        try:
            with transaction.atomic():
                Ficha.objects.create(
                    idficha=siguiente,
                    num_ficha=num_ficha,
                    instructor_usuario_id_usuario_id=inst_id,
                )
        except IntegrityError:
            messages.error(request, "No se pudo crear la ficha. Revise número e instructor.")
            return _render_admin(
                request,
                "formFicha.html",
                {"instructores": instructores, "ficha": {"numFicha": num_raw, "instructorId": str(inst_id)}},
            )

        messages.success(request, "Ficha creada correctamente.")
        return redirect("listar_fichas")

    return _render_admin(
        request,
        "formFicha.html",
        {"instructores": instructores, "ficha": {**EMPTY_FICHA.copy(), "instructorId": ""}},
    )


def editar_ficha(request, ficha_id):
    f = get_object_or_404(
        Ficha.objects.select_related("instructor_usuario_id_usuario"),
        pk=ficha_id,
    )
    instructores = _instructores_para_select()
    ficha_ctx = {
        "numFicha": str(f.num_ficha),
        "instructorId": str(f.instructor_usuario_id_usuario_id),
    }

    if request.method == "POST":
        num_raw = (request.POST.get("numFicha") or "").strip()
        inst_raw = request.POST.get("instructorId")
        try:
            num_ficha = int(num_raw)
        except ValueError:
            messages.error(request, "Número de ficha inválido.")
            return _render_admin(
                request,
                "formFicha.html",
                {
                    "instructores": instructores,
                    "ficha": {"numFicha": num_raw, "instructorId": inst_raw or ficha_ctx["instructorId"]},
                    "idFichaEditar": ficha_id,
                },
            )
        if num_ficha < 1 or num_ficha > MAX_MEDIUMINT:
            messages.error(
                request,
                f"El número de ficha debe estar entre 1 y {MAX_MEDIUMINT}.",
            )
            return _render_admin(
                request,
                "formFicha.html",
                {
                    "instructores": instructores,
                    "ficha": {"numFicha": num_raw, "instructorId": inst_raw or ficha_ctx["instructorId"]},
                    "idFichaEditar": ficha_id,
                },
            )
        if not inst_raw:
            messages.error(request, "Seleccione un instructor.")
            return _render_admin(
                request,
                "formFicha.html",
                {"instructores": instructores, "ficha": ficha_ctx, "idFichaEditar": ficha_id},
            )
        try:
            inst_id = int(inst_raw)
        except ValueError:
            messages.error(request, "Instructor no válido.")
            return _render_admin(
                request,
                "formFicha.html",
                {"instructores": instructores, "ficha": ficha_ctx, "idFichaEditar": ficha_id},
            )
        if not Instructor.objects.filter(pk=inst_id).exists():
            messages.error(request, "El instructor elegido no existe.")
            return _render_admin(
                request,
                "formFicha.html",
                {"instructores": instructores, "ficha": ficha_ctx, "idFichaEditar": ficha_id},
            )

        try:
            actualizados = Ficha.objects.filter(pk=f.pk).update(
                num_ficha=num_ficha,
                instructor_usuario_id_usuario_id=inst_id,
            )
        except IntegrityError:
            messages.error(request, "No se pudo guardar la ficha.")
            return _render_admin(
                request,
                "formFicha.html",
                {"instructores": instructores, "ficha": ficha_ctx, "idFichaEditar": ficha_id},
            )
        if not actualizados:
            messages.error(request, "No se pudo actualizar la ficha.")
            return _render_admin(
                request,
                "formFicha.html",
                {"instructores": instructores, "ficha": ficha_ctx, "idFichaEditar": ficha_id},
            )

        messages.success(request, "Ficha actualizada correctamente.")
        return redirect("listar_fichas")

    return _render_admin(
        request,
        "formFicha.html",
        {
            "instructores": instructores,
            "ficha": ficha_ctx,
            "idFichaEditar": ficha_id,
        },
    )


@require_POST
def eliminar_ficha(request, ficha_id):
    f = Ficha.objects.filter(pk=ficha_id).first()
    if not f:
        messages.error(request, "Ficha no encontrada.")
        return redirect("listar_fichas")
    try:
        f.delete()
    except IntegrityError:
        messages.error(
            request,
            "No se puede eliminar: hay aprendices u otros registros asociados a esta ficha.",
        )
        return redirect("listar_fichas")
    messages.success(request, "Ficha eliminada correctamente.")
    return redirect("listar_fichas")


def _programa_form_desde_modelo(p: Programas):
    return {
        "nombrePrograma": p.nombre_programa,
        "nivelFormacion": p.nivel_formacion,
        "duracion": p.duracion,
        "jornadaId": str(p.jornada_id),
        "modalidadId": str(p.modalidad_id),
        "coordinacionId": str(p.coordinacion_id),
    }


def _recurso_form_desde_modelo(r: Recursos):
    return {
        "nombre": r.nombre_recurso,
        "serial": r.serial_recurso,
        "numero": r.num_recurso,
        "idTipoRecurso": str(r.tipo_recurso_id),
        "idAmbiente": str(r.ambiente_id),
        "estado": (r.estado or "").strip(),
        "observacion": r.observacion or "",
    }


def _ambiente_form_desde_modelo(a: Ambiente):
    return {
        "numero": a.num_ambiente,
        "capacidad": a.capacidad,
        "tipo": a.tipo_ambiente,
        "estado": a.estado,
    }


def listar_programas(request):
    q = admin_q(request)
    qs = Programas.objects.select_related("jornada", "modalidad", "coordinacion").order_by(
        "id_programas"
    )
    if q:
        nu = maybe_int(q)
        cond = (
            Q(nombre_programa__icontains=q)
            | Q(nivel_formacion__icontains=q)
            | Q(duracion__icontains=q)
            | Q(jornada__nombre_jornada__icontains=q)
            | Q(modalidad__nombre_modalidad__icontains=q)
            | Q(coordinacion__nombre_coordinacion__icontains=q)
        )
        if nu is not None:
            cond |= Q(id_programas=nu)
        qs = qs.filter(cond)

    filas = []
    for p in qs:
        filas.append(
            {
                "p_id": p.id_programas,
                "p_nombre": (p.nombre_programa or "").strip() or "—",
                "p_nivel": (p.nivel_formacion or "").strip() or "—",
                "p_duracion": (p.duracion or "").strip() or "—",
                "p_jornada": (p.jornada.nombre_jornada or "").strip() or "—",
                "p_modalidad": (p.modalidad.nombre_modalidad or "").strip() or "—",
                "p_coordinacion": (p.coordinacion.nombre_coordinacion or "").strip() or "—",
            }
        )
    return _render_admin(request, "listarProgramas.html", {"programas": filas, "q": q})


def _context_form_programa(programa_dict=None):
    prog = EMPTY_PROGRAMA.copy()
    if programa_dict:
        prog.update(programa_dict)
    return {
        "programa": prog,
        "jornadas": list(Jornada.objects.all().order_by("id_jornada")),
        "modalidades": list(Modalidad.objects.all().order_by("id_modalidad")),
        "coordinaciones": list(Coordinacion.objects.all().order_by("id_coordinacion")),
    }


def crear_programa(request):
    ctx = _context_form_programa()
    if request.method == "POST":
        nombre = request.POST.get("nombrePrograma", "").strip()
        nivel = request.POST.get("nivelFormacion", "").strip()
        duracion = request.POST.get("duracion", "").strip()
        jid = request.POST.get("jornadaId")
        mid = request.POST.get("modalidadId")
        cid = request.POST.get("coordinacionId")
        if not all([nombre, nivel, duracion, jid, mid, cid]):
            messages.error(request, "Complete todos los campos.")
            ctx["programa"].update(
                {
                    "nombrePrograma": nombre,
                    "nivelFormacion": nivel,
                    "duracion": duracion,
                    "jornadaId": jid or "",
                    "modalidadId": mid or "",
                    "coordinacionId": cid or "",
                }
            )
            return _render_admin(request, "formPrograma.html", ctx)
        try:
            jid, mid, cid = int(jid), int(mid), int(cid)
        except ValueError:
            messages.error(request, "Valores de jornada, modalidad o coordinación inválidos.")
            return _render_admin(request, "formPrograma.html", ctx)
        if not (
            Jornada.objects.filter(pk=jid).exists()
            and Modalidad.objects.filter(pk=mid).exists()
            and Coordinacion.objects.filter(pk=cid).exists()
        ):
            messages.error(request, "La jornada, modalidad o coordinación no existe.")
            return _render_admin(request, "formPrograma.html", ctx)
        siguiente = (Programas.objects.aggregate(m=Max("id_programas"))["m"] or 0) + 1
        try:
            Programas.objects.create(
                id_programas=siguiente,
                nombre_programa=nombre,
                nivel_formacion=nivel,
                duracion=duracion,
                jornada_id=jid,
                modalidad_id=mid,
                coordinacion_id=cid,
            )
        except IntegrityError:
            messages.error(request, "No se pudo crear el programa.")
            return _render_admin(request, "formPrograma.html", ctx)
        messages.success(request, "Programa creado correctamente.")
        return redirect("listar_programas")
    return _render_admin(request, "formPrograma.html", ctx)


def editar_programa(request, programa_id):
    p = get_object_or_404(Programas, pk=programa_id)
    ctx = _context_form_programa(_programa_form_desde_modelo(p))
    ctx["idProgramaEditar"] = programa_id
    if request.method == "POST":
        nombre = request.POST.get("nombrePrograma", "").strip()
        nivel = request.POST.get("nivelFormacion", "").strip()
        duracion = request.POST.get("duracion", "").strip()
        jid = request.POST.get("jornadaId")
        mid = request.POST.get("modalidadId")
        cid = request.POST.get("coordinacionId")
        if not all([nombre, nivel, duracion, jid, mid, cid]):
            messages.error(request, "Complete todos los campos.")
            ctx["programa"].update(
                {
                    "nombrePrograma": nombre,
                    "nivelFormacion": nivel,
                    "duracion": duracion,
                    "jornadaId": jid or "",
                    "modalidadId": mid or "",
                    "coordinacionId": cid or "",
                }
            )
            return _render_admin(request, "formPrograma.html", ctx)
        try:
            jid, mid, cid = int(jid), int(mid), int(cid)
        except ValueError:
            messages.error(request, "Valores inválidos.")
            return _render_admin(request, "formPrograma.html", ctx)
        p.nombre_programa = nombre
        p.nivel_formacion = nivel
        p.duracion = duracion
        p.jornada_id = jid
        p.modalidad_id = mid
        p.coordinacion_id = cid
        try:
            p.save()
        except IntegrityError:
            messages.error(request, "No se pudo guardar el programa.")
            return _render_admin(request, "formPrograma.html", ctx)
        messages.success(request, "Programa actualizado correctamente.")
        return redirect("listar_programas")
    return _render_admin(request, "formPrograma.html", ctx)


@require_POST
def eliminar_programa(request, programa_id):
    p = Programas.objects.filter(pk=programa_id).first()
    if not p:
        messages.error(request, "Programa no encontrado.")
        return redirect("listar_programas")
    try:
        p.delete()
    except IntegrityError:
        messages.error(
            request,
            "No se puede eliminar: hay aprendices u otros registros que usan este programa.",
        )
        return redirect("listar_programas")
    messages.success(request, "Programa eliminado correctamente.")
    return redirect("listar_programas")


def listar_recursos(request):
    q = admin_q(request)
    filtro = Q()
    if q:
        nu = maybe_int(q)
        filtro = (
            Q(nombre_recurso__icontains=q)
            | Q(serial_recurso__icontains=q)
            | Q(estado__icontains=q)
            | Q(observacion__icontains=q)
            | Q(tipo_recurso__recurso_tipo__icontains=q)
        )
        if nu is not None:
            filtro |= Q(id_recurso=nu) | Q(num_recurso=nu) | Q(ambiente__num_ambiente=nu)
    filas = []
    qs = (
        Recursos.objects.filter(filtro)
        .select_related("tipo_recurso", "ambiente")
        .order_by("id_recurso")
    )
    for r in qs:
        filas.append(
            {
                "r_id": r.id_recurso,
                "r_nombre": (r.nombre_recurso or "").strip() or "—",
                "r_serial": (r.serial_recurso or "").strip() or "—",
                "r_numero": str(r.num_recurso),
                "r_tipo": (r.tipo_recurso.recurso_tipo or "").strip() or "—",
                "r_estado": (r.estado or "").strip() or "—",
                "r_observacion": (r.observacion or "").strip(),
                "r_ambiente": str(r.ambiente.num_ambiente),
            }
        )
    return _render_admin(request, "listarRecursos.html", {"recursos": filas, "q": q})


def _context_form_recurso(recurso_dict=None):
    rec = EMPTY_RESOURCE.copy()
    if recurso_dict:
        rec.update(recurso_dict)
    return {
        "recurso": rec,
        "tipos": list(TipoRecurso.objects.all().order_by("id_tipo_recurso")),
        "ambientes": list(Ambiente.objects.all().order_by("id_ambiente")),
    }


def crear_recurso(request):
    ctx = _context_form_recurso()
    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        serial = request.POST.get("serial", "").strip()
        estado = request.POST.get("estado", "").strip()
        observacion = (request.POST.get("observacion") or "").strip()
        tipo_id = request.POST.get("idTipoRecurso")
        amb_id = request.POST.get("idAmbiente")
        try:
            numero = int(request.POST.get("numero"))
        except (TypeError, ValueError):
            messages.error(request, "El número del recurso debe ser entero.")
            ctx["recurso"].update(
                {
                    "nombre": nombre,
                    "serial": serial,
                    "numero": request.POST.get("numero") or "",
                    "idTipoRecurso": tipo_id or "",
                    "idAmbiente": amb_id or "",
                    "estado": estado,
                    "observacion": observacion,
                }
            )
            return _render_admin(request, "formRecurso.html", ctx)
        if not all([nombre, serial, tipo_id, amb_id, estado]):
            messages.error(request, "Complete los campos obligatorios.")
            ctx["recurso"].update(
                {
                    "nombre": nombre,
                    "serial": serial,
                    "numero": numero,
                    "idTipoRecurso": tipo_id or "",
                    "idAmbiente": amb_id or "",
                    "estado": estado,
                    "observacion": observacion,
                }
            )
            return _render_admin(request, "formRecurso.html", ctx)
        try:
            tipo_id, amb_id = int(tipo_id), int(amb_id)
        except ValueError:
            messages.error(request, "Tipo o ambiente no válido.")
            return _render_admin(request, "formRecurso.html", ctx)
        if not TipoRecurso.objects.filter(pk=tipo_id).exists() or not Ambiente.objects.filter(
            pk=amb_id
        ).exists():
            messages.error(request, "Tipo de recurso o ambiente no existe.")
            return _render_admin(request, "formRecurso.html", ctx)
        try:
            Recursos.objects.create(
                nombre_recurso=nombre,
                serial_recurso=serial,
                num_recurso=numero,
                tipo_recurso_id=tipo_id,
                ambiente_id=amb_id,
                estado=estado or None,
                observacion=observacion or None,
            )
        except IntegrityError:
            messages.error(request, "No se pudo crear el recurso.")
            return _render_admin(request, "formRecurso.html", ctx)
        messages.success(request, "Recurso creado correctamente.")
        return redirect("listar_recursos")
    return _render_admin(request, "formRecurso.html", ctx)


def editar_recurso(request, recurso_id):
    r = get_object_or_404(Recursos, pk=recurso_id)
    ctx = _context_form_recurso(_recurso_form_desde_modelo(r))
    ctx["idRecursoEditar"] = recurso_id
    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        serial = request.POST.get("serial", "").strip()
        estado = request.POST.get("estado", "").strip()
        observacion = (request.POST.get("observacion") or "").strip()
        tipo_id = request.POST.get("idTipoRecurso")
        amb_id = request.POST.get("idAmbiente")
        try:
            numero = int(request.POST.get("numero"))
        except (TypeError, ValueError):
            messages.error(request, "El número del recurso debe ser entero.")
            return _render_admin(request, "formRecurso.html", ctx)
        if not all([nombre, serial, tipo_id, amb_id, estado]):
            messages.error(request, "Complete los campos obligatorios.")
            return _render_admin(request, "formRecurso.html", ctx)
        try:
            tipo_id, amb_id = int(tipo_id), int(amb_id)
        except ValueError:
            messages.error(request, "Tipo o ambiente no válido.")
            return _render_admin(request, "formRecurso.html", ctx)
        r.nombre_recurso = nombre
        r.serial_recurso = serial
        r.num_recurso = numero
        r.tipo_recurso_id = tipo_id
        r.ambiente_id = amb_id
        r.estado = estado or None
        r.observacion = observacion or None
        try:
            r.save()
        except IntegrityError:
            messages.error(request, "No se pudo guardar el recurso.")
            return _render_admin(request, "formRecurso.html", ctx)
        messages.success(request, "Recurso actualizado correctamente.")
        return redirect("listar_recursos")
    return _render_admin(request, "formRecurso.html", ctx)


@require_POST
def eliminar_recurso(request, recurso_id):
    r = Recursos.objects.filter(pk=recurso_id).first()
    if not r:
        messages.error(request, "Recurso no encontrado.")
        return redirect("listar_recursos")
    try:
        r.delete()
    except IntegrityError:
        messages.error(request, "No se puede eliminar: el recurso está referenciado en otros registros.")
        return redirect("listar_recursos")
    messages.success(request, "Recurso eliminado correctamente.")
    return redirect("listar_recursos")


def listar_ambientes(request):
    q = admin_q(request)
    qs = Ambiente.objects.all().order_by("id_ambiente")
    if q:
        nu = maybe_int(q)
        cond = Q(tipo_ambiente__icontains=q) | Q(estado__icontains=q)
        if nu is not None:
            cond |= Q(id_ambiente=nu) | Q(num_ambiente=nu) | Q(capacidad=nu)
        qs = qs.filter(cond)

    filas = []
    for am in qs:
        filas.append(
            {
                "am_id": am.id_ambiente,
                "am_num": str(am.num_ambiente),
                "am_capaci": str(am.capacidad),
                "am_tipo": (am.tipo_ambiente or "").strip() or "—",
                "am_estado": (am.estado or "").strip() or "—",
            }
        )
    return _render_admin(request, "listarAmbientes.html", {"ambientes": filas, "q": q})


def _context_form_ambiente(ambiente_dict=None):
    amb = EMPTY_AMBIENTE.copy()
    if ambiente_dict:
        amb.update(ambiente_dict)
    return {"ambiente": amb}


def crear_ambiente(request):
    ctx = _context_form_ambiente()
    if request.method == "POST":
        try:
            numero = int(request.POST.get("numero"))
            capacidad = int(request.POST.get("capacidad"))
        except (TypeError, ValueError):
            messages.error(request, "Número y capacidad deben ser enteros.")
            ctx["ambiente"].update(
                {
                    "numero": request.POST.get("numero") or "",
                    "capacidad": request.POST.get("capacidad") or "",
                    "tipo": request.POST.get("tipo", "").strip(),
                    "estado": request.POST.get("estado", "").strip(),
                }
            )
            return _render_admin(request, "formAmbiente.html", ctx)
        tipo = request.POST.get("tipo", "").strip()
        estado = request.POST.get("estado", "").strip()
        if not tipo or not estado:
            messages.error(request, "Tipo y estado son obligatorios.")
            ctx["ambiente"].update(
                {"numero": numero, "capacidad": capacidad, "tipo": tipo, "estado": estado}
            )
            return _render_admin(request, "formAmbiente.html", ctx)
        siguiente = (Ambiente.objects.aggregate(m=Max("id_ambiente"))["m"] or 0) + 1
        try:
            Ambiente.objects.create(
                id_ambiente=siguiente,
                num_ambiente=numero,
                capacidad=capacidad,
                tipo_ambiente=tipo,
                estado=estado,
            )
        except IntegrityError:
            messages.error(request, "No se pudo crear el ambiente.")
            return _render_admin(request, "formAmbiente.html", ctx)
        messages.success(request, "Ambiente creado correctamente.")
        return redirect("listar_ambientes")
    return _render_admin(request, "formAmbiente.html", ctx)


def editar_ambiente(request, ambiente_id):
    a = get_object_or_404(Ambiente, pk=ambiente_id)
    ctx = _context_form_ambiente(_ambiente_form_desde_modelo(a))
    ctx["idAmbienteEditar"] = ambiente_id
    if request.method == "POST":
        try:
            numero = int(request.POST.get("numero"))
            capacidad = int(request.POST.get("capacidad"))
        except (TypeError, ValueError):
            messages.error(request, "Número y capacidad deben ser enteros.")
            return _render_admin(request, "formAmbiente.html", ctx)
        tipo = request.POST.get("tipo", "").strip()
        estado = request.POST.get("estado", "").strip()
        if not tipo or not estado:
            messages.error(request, "Tipo y estado son obligatorios.")
            return _render_admin(request, "formAmbiente.html", ctx)
        a.num_ambiente = numero
        a.capacidad = capacidad
        a.tipo_ambiente = tipo
        a.estado = estado
        try:
            a.save()
        except IntegrityError:
            messages.error(request, "No se pudo guardar el ambiente.")
            return _render_admin(request, "formAmbiente.html", ctx)
        messages.success(request, "Ambiente actualizado correctamente.")
        return redirect("listar_ambientes")
    return _render_admin(request, "formAmbiente.html", ctx)


@require_POST
def eliminar_ambiente(request, ambiente_id):
    a = Ambiente.objects.filter(pk=ambiente_id).first()
    if not a:
        messages.error(request, "Ambiente no encontrado.")
        return redirect("listar_ambientes")
    try:
        a.delete()
    except IntegrityError:
        messages.error(
            request,
            "No se puede eliminar: hay recursos, incidentes u otros registros en este ambiente.",
        )
        return redirect("listar_ambientes")
    messages.success(request, "Ambiente eliminado correctamente.")
    return redirect("listar_ambientes")



















































































































def _admin_excel_headers(ws, headers):
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def exportar_usuarios_excel(request):
    q = admin_q(request)
    roles_por_usuario = defaultdict(list)
    for ur in UserRol.objects.select_related("id_rol").all():
        roles_por_usuario[ur.id_usuario_id].append(ur.id_rol.nombre_rol)
    qs = Usuario.objects.all().order_by("id_usuario")
    if q:
        nu = maybe_int(q)
        cond = (
            Q(p_nombre__icontains=q)
            | Q(s_nombre__icontains=q)
            | Q(p_apellido__icontains=q)
            | Q(s_apellido__icontains=q)
            | Q(correo__icontains=q)
            | Q(tipo_documento__icontains=q)
            | Q(userrol_set__id_rol__nombre_rol__icontains=q)
        )
        if nu is not None:
            cond |= Q(id_usuario=nu) | Q(num_documento=nu)
        qs = qs.filter(cond).distinct()
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    _admin_excel_headers(ws, ["ID", "Nombre", "Apellido", "Correo", "Rol"])
    for u in qs:
        nombres = " ".join(filter(None, [u.p_nombre, u.s_nombre])).strip()
        apellidos = " ".join(filter(None, [u.p_apellido, u.s_apellido])).strip()
        roles = ", ".join(roles_por_usuario.get(u.id_usuario, [])) or "—"
        ws.append([u.id_usuario, nombres or "—", apellidos or "—", u.correo, roles])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="usuarios_{admin_safe_q(q)}.xlsx"'
    return resp


def exportar_usuarios_pdf(request):
    q = admin_q(request)
    roles_por_usuario = defaultdict(list)
    for ur in UserRol.objects.select_related("id_rol").all():
        roles_por_usuario[ur.id_usuario_id].append(ur.id_rol.nombre_rol)
    qs = Usuario.objects.all().order_by("id_usuario")
    if q:
        nu = maybe_int(q)
        cond = (
            Q(p_nombre__icontains=q)
            | Q(s_nombre__icontains=q)
            | Q(p_apellido__icontains=q)
            | Q(s_apellido__icontains=q)
            | Q(correo__icontains=q)
            | Q(tipo_documento__icontains=q)
            | Q(userrol_set__id_rol__nombre_rol__icontains=q)
        )
        if nu is not None:
            cond |= Q(id_usuario=nu) | Q(num_documento=nu)
        qs = qs.filter(cond).distinct()
    rows = []
    for u in qs:
        nombres = " ".join(filter(None, [u.p_nombre, u.s_nombre])).strip()
        apellidos = " ".join(filter(None, [u.p_apellido, u.s_apellido])).strip()
        roles = ", ".join(roles_por_usuario.get(u.id_usuario, [])) or "—"
        rows.append([str(u.id_usuario), nombres or "—", apellidos or "—", u.correo, roles])
    return landscape_pdf_response(
        "Reporte de Usuarios",
        q,
        ["ID", "Nombre", "Apellido", "Correo", "Rol"],
        rows,
        [0.08, 0.22, 0.22, 0.28, 0.20],
        f"usuarios_{admin_safe_q(q)}.pdf",
    )


def exportar_fichas_excel(request):
    q = admin_q(request)
    qs = Ficha.objects.select_related("instructor_usuario_id_usuario__usuario_id_usuario").order_by(
        "num_ficha"
    )
    if q:
        nu = maybe_int(q)
        uq = (
            Q(instructor_usuario_id_usuario__usuario_id_usuario__p_nombre__icontains=q)
            | Q(instructor_usuario_id_usuario__usuario_id_usuario__s_nombre__icontains=q)
            | Q(instructor_usuario_id_usuario__usuario_id_usuario__p_apellido__icontains=q)
            | Q(instructor_usuario_id_usuario__usuario_id_usuario__s_apellido__icontains=q)
        )
        if nu is not None:
            uq |= Q(idficha=nu) | Q(num_ficha=nu)
        qs = qs.filter(uq).distinct()
    wb = Workbook()
    ws = wb.active
    ws.title = "Fichas"
    _admin_excel_headers(ws, ["ID", "Número ficha", "Instructor"])
    for f in qs:
        u = f.instructor_usuario_id_usuario.usuario_id_usuario
        instructor = " ".join(filter(None, [u.p_nombre, u.p_apellido])).strip() or "—"
        ws.append([f.idficha, f.num_ficha, instructor])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="fichas_{admin_safe_q(q)}.xlsx"'
    return resp


def exportar_fichas_pdf(request):
    q = admin_q(request)
    qs = Ficha.objects.select_related("instructor_usuario_id_usuario__usuario_id_usuario").order_by(
        "num_ficha"
    )
    if q:
        nu = maybe_int(q)
        uq = (
            Q(instructor_usuario_id_usuario__usuario_id_usuario__p_nombre__icontains=q)
            | Q(instructor_usuario_id_usuario__usuario_id_usuario__s_nombre__icontains=q)
            | Q(instructor_usuario_id_usuario__usuario_id_usuario__p_apellido__icontains=q)
            | Q(instructor_usuario_id_usuario__usuario_id_usuario__s_apellido__icontains=q)
        )
        if nu is not None:
            uq |= Q(idficha=nu) | Q(num_ficha=nu)
        qs = qs.filter(uq).distinct()
    rows = []
    for f in qs:
        u = f.instructor_usuario_id_usuario.usuario_id_usuario
        instructor = " ".join(filter(None, [u.p_nombre, u.p_apellido])).strip() or "—"
        rows.append([str(f.idficha), str(f.num_ficha), instructor])
    return landscape_pdf_response(
        "Reporte de Fichas",
        q,
        ["ID", "Número ficha", "Instructor"],
        rows,
        [0.12, 0.28, 0.60],
        f"fichas_{admin_safe_q(q)}.pdf",
    )


def exportar_programas_excel(request):
    q = admin_q(request)
    qs = Programas.objects.select_related("jornada", "modalidad", "coordinacion").order_by("id_programas")
    if q:
        nu = maybe_int(q)
        cond = (
            Q(nombre_programa__icontains=q)
            | Q(nivel_formacion__icontains=q)
            | Q(duracion__icontains=q)
            | Q(jornada__nombre_jornada__icontains=q)
            | Q(modalidad__nombre_modalidad__icontains=q)
            | Q(coordinacion__nombre_coordinacion__icontains=q)
        )
        if nu is not None:
            cond |= Q(id_programas=nu)
        qs = qs.filter(cond)
    wb = Workbook()
    ws = wb.active
    ws.title = "Programas"
    _admin_excel_headers(
        ws,
        ["ID", "Nombre", "Nivel", "Duración", "Jornada", "Modalidad", "Coordinación"],
    )
    for p in qs:
        ws.append(
            [
                p.id_programas,
                (p.nombre_programa or "").strip() or "—",
                (p.nivel_formacion or "").strip() or "—",
                (p.duracion or "").strip() or "—",
                (p.jornada.nombre_jornada or "").strip() or "—",
                (p.modalidad.nombre_modalidad or "").strip() or "—",
                (p.coordinacion.nombre_coordinacion or "").strip() or "—",
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="programas_{admin_safe_q(q)}.xlsx"'
    return resp


def exportar_programas_pdf(request):
    q = admin_q(request)
    qs = Programas.objects.select_related("jornada", "modalidad", "coordinacion").order_by("id_programas")
    if q:
        nu = maybe_int(q)
        cond = (
            Q(nombre_programa__icontains=q)
            | Q(nivel_formacion__icontains=q)
            | Q(duracion__icontains=q)
            | Q(jornada__nombre_jornada__icontains=q)
            | Q(modalidad__nombre_modalidad__icontains=q)
            | Q(coordinacion__nombre_coordinacion__icontains=q)
        )
        if nu is not None:
            cond |= Q(id_programas=nu)
        qs = qs.filter(cond)
    rows = []
    for p in qs:
        rows.append(
            [
                str(p.id_programas),
                (p.nombre_programa or "").strip() or "—",
                (p.nivel_formacion or "").strip() or "—",
                (p.duracion or "").strip() or "—",
                (p.jornada.nombre_jornada or "").strip() or "—",
                (p.modalidad.nombre_modalidad or "").strip() or "—",
                (p.coordinacion.nombre_coordinacion or "").strip() or "—",
            ]
        )
    return landscape_pdf_response(
        "Reporte de Programas",
        q,
        ["ID", "Nombre", "Nivel", "Duración", "Jornada", "Modalidad", "Coordinación"],
        rows,
        [0.06, 0.20, 0.12, 0.10, 0.14, 0.14, 0.24],
        f"programas_{admin_safe_q(q)}.pdf",
    )


def exportar_recursos_excel(request):
    q = admin_q(request)
    filtro = Q()
    if q:
        nu = maybe_int(q)
        filtro = (
            Q(nombre_recurso__icontains=q)
            | Q(serial_recurso__icontains=q)
            | Q(estado__icontains=q)
            | Q(observacion__icontains=q)
            | Q(tipo_recurso__recurso_tipo__icontains=q)
        )
        if nu is not None:
            filtro |= Q(id_recurso=nu) | Q(num_recurso=nu) | Q(ambiente__num_ambiente=nu)
    qs = Recursos.objects.filter(filtro).select_related("tipo_recurso", "ambiente").order_by("id_recurso")
    wb = Workbook()
    ws = wb.active
    ws.title = "Recursos"
    _admin_excel_headers(
        ws,
        ["ID", "Nombre", "Serial", "Número", "Tipo", "Estado", "Ambiente", "Observación"],
    )
    for r in qs:
        ws.append(
            [
                r.id_recurso,
                (r.nombre_recurso or "").strip() or "—",
                (r.serial_recurso or "").strip() or "—",
                r.num_recurso,
                (r.tipo_recurso.recurso_tipo or "").strip() or "—",
                (r.estado or "").strip() or "—",
                str(r.ambiente.num_ambiente) if r.ambiente_id else "",
                (r.observacion or "").strip(),
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="recursos_{admin_safe_q(q)}.xlsx"'
    return resp


def exportar_recursos_pdf(request):
    q = admin_q(request)
    filtro = Q()
    if q:
        nu = maybe_int(q)
        filtro = (
            Q(nombre_recurso__icontains=q)
            | Q(serial_recurso__icontains=q)
            | Q(estado__icontains=q)
            | Q(observacion__icontains=q)
            | Q(tipo_recurso__recurso_tipo__icontains=q)
        )
        if nu is not None:
            filtro |= Q(id_recurso=nu) | Q(num_recurso=nu) | Q(ambiente__num_ambiente=nu)
    qs = Recursos.objects.filter(filtro).select_related("tipo_recurso", "ambiente").order_by("id_recurso")
    rows = []
    for r in qs:
        rows.append(
            [
                str(r.id_recurso),
                (r.nombre_recurso or "").strip() or "—",
                (r.serial_recurso or "").strip() or "—",
                str(r.num_recurso),
                (r.tipo_recurso.recurso_tipo or "").strip() or "—",
                (r.estado or "").strip() or "—",
                str(r.ambiente.num_ambiente) if r.ambiente_id else "",
                (r.observacion or "").strip(),
            ]
        )
    return landscape_pdf_response(
        "Reporte de Recursos",
        q,
        ["ID", "Nombre", "Serial", "Número", "Tipo", "Estado", "Ambiente", "Observación"],
        rows,
        [0.06, 0.14, 0.12, 0.07, 0.12, 0.10, 0.09, 0.30],
        f"recursos_{admin_safe_q(q)}.pdf",
    )


def exportar_ambientes_admin_excel(request):
    q = admin_q(request)
    qs = Ambiente.objects.all().order_by("id_ambiente")
    if q:
        nu = maybe_int(q)
        cond = Q(tipo_ambiente__icontains=q) | Q(estado__icontains=q)
        if nu is not None:
            cond |= Q(id_ambiente=nu) | Q(num_ambiente=nu) | Q(capacidad=nu)
        qs = qs.filter(cond)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ambientes"
    _admin_excel_headers(ws, ["ID", "Número", "Capacidad", "Tipo", "Estado"])
    for a in qs:
        ws.append(
            [
                a.id_ambiente,
                a.num_ambiente,
                a.capacidad,
                (a.tipo_ambiente or "").strip(),
                (a.estado or "").strip(),
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="ambientes_{admin_safe_q(q)}.xlsx"'
    return resp


def exportar_ambientes_admin_pdf(request):
    q = admin_q(request)
    qs = Ambiente.objects.all().order_by("id_ambiente")
    if q:
        nu = maybe_int(q)
        cond = Q(tipo_ambiente__icontains=q) | Q(estado__icontains=q)
        if nu is not None:
            cond |= Q(id_ambiente=nu) | Q(num_ambiente=nu) | Q(capacidad=nu)
        qs = qs.filter(cond)
    rows = []
    for a in qs:
        rows.append(
            [
                str(a.id_ambiente),
                str(a.num_ambiente),
                str(a.capacidad),
                (a.tipo_ambiente or "").strip(),
                (a.estado or "").strip(),
            ]
        )
    return landscape_pdf_response(
        "Reporte de Ambientes",
        q,
        ["ID", "Número", "Capacidad", "Tipo", "Estado"],
        rows,
        [0.12, 0.14, 0.14, 0.35, 0.25],
        f"ambientes_{admin_safe_q(q)}.pdf",
    )


##⣿⣿⣿⣿⣇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠈⠉⠛⠻⣿⣿⣿⣿⣿⣿
##⣿⣿⣿⣿⣿⣦⠀⠀⠀⠀⠀⠀⠀⠀⢀⣤⣄⡀⠀⢻⣿⣿⣿⣿⣿
##⣿⣿⣿⣿⣿⣿⣇⠀⠀⠀⠀⠀⠀⠀⠸⣿⣿⣿⠃⢰⣿⣿⣿⣿⣿
##⣿⣿⣿⣿⣿⣿⣿⣆⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢀⣼⣿⣿⣿⣿⣿
##⣿⣿⣿⣿⣿⣿⣿⣿⡆⠀⠀⠀⠀⠀⠀⢶⣶⣶⣾⣿⣿⣿⣿⣿⣿
##⣿⣿⣿⣿⣿⣿⣿⣿⣧⠀⢠⡀⠐⠀⠀⠀⠻⢿⣿⣿⣿⣿⣿⣿⣿
##⣿⣿⣿⣿⣿⣿⣿⣿⣿⡄⢸⣷⡄⠀⠣⣄⡀⠀⠉⠛⢿⣿⣿⣿⣿
##⣿⣿⣿⣿⣿⣿⣿⣿⣿⣇⠀⣿⣿⣦⠀⠹⣿⣷⣶⣦⣼⣿⣿⣿⣿
##⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣼⣿⣿⣿⣷⣄⣸⣿⣿⣿⣿⣿⣿⣿⣿ Mas bien siga revisando el codigo menor
## easter egg 1/10